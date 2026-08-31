"""Fachada del paquete de parseo.

**La ruta `importaciones.parsing` no se puede romper**: la importan seis
módulos, incluidas dos migraciones históricas
(`rutinas/migrations/0006_backfill_grupo_muscular_snapshot.py` y
`ejercicios/migrations/0003_backfill_categorias.py`). Por eso este módulo
re-exporta la API pública entera y no tiene lógica propia salvo el
despachador de layout de `leer_hoja_plantilla`.
"""

import openpyxl

from importaciones.parsing.comun import (  # noqa: F401  (re-export)
    ALIAS_BIBLIOTECA,
    FILAS_BUSQUEDA_ENCABEZADO,
    ALIAS_PLANTILLA,
    ColumnaRequeridaFaltante,
    FilaInvalida,
    HojaParseada,
    ItemParseado,
    buscar_fila_encabezado,
    detectar_columnas,
    mejor_encabezado_parcial,
    normalizar_texto,
)
from importaciones.parsing.ancha import (  # noqa: F401  (re-export)
    detectar_matriz_ancha,
    leer_hoja_ancha,
)
from importaciones.parsing.tabular import leer_hoja_biblioteca, leer_hoja_larga


def leer_hoja_plantilla(ws):
    """Único punto donde se elige el layout de una hoja de PLANTILLAS.

    La matriz ancha se prueba PRIMERO, siempre: si se probara al revés, una
    hoja ancha matchearía igual el layout largo (su fila de grupos tiene
    "EJERCICIOS" y la de subcampos tiene "Series"/"Reps"/"Carga") y produciría
    filas plausibles con las columnas corridas. Basura silenciosa es peor que
    cero items. Al revés no puede pasar: `RE_SEMANA` exige el dígito, así que
    el "Semana" a secas del layout largo no matchea nunca.
    """
    encabezado_ancho = detectar_matriz_ancha(ws)
    if encabezado_ancho is not None:
        return leer_hoja_ancha(ws, encabezado_ancho)
    return leer_hoja_larga(ws)


def parsear_archivo_plantillas(archivo):
    """Abre `archivo` (un `UploadedFile` de Django) y devuelve una
    `HojaParseada` por cada hoja del workbook (decisión 7 del spec:
    multi-hoja -> multi-plantilla)."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    return [leer_hoja_plantilla(wb[nombre]) for nombre in wb.sheetnames]


def parsear_archivo_biblioteca(archivo):
    """El import de biblioteca usa solo la primera hoja del archivo."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    return leer_hoja_biblioteca(wb[wb.sheetnames[0]])
