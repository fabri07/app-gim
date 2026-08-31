"""Tests de `importaciones`. Ver `rutinas/tests.py` para el estilo de
fixtures de este repo."""

import io
import json

import openpyxl
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.test import SimpleTestCase, TestCase
from django.urls import reverse

from ejercicios.models import CategoriaEjercicio, Ejercicio
from importaciones.matching import (
    resolver_categorias,
    MatchResultado,
    construir_indice_ejercicios,
    resolver_nombre,
)
from importaciones.models import Importacion
from importaciones.parsing import (
    ColumnaRequeridaFaltante,
    ALIAS_BIBLIOTECA,
    ALIAS_PLANTILLA,
    FilaInvalida,
    HojaParseada,
    ItemParseado,
    buscar_fila_encabezado,
    detectar_columnas,
    detectar_matriz_ancha,
    mejor_encabezado_parcial,
    leer_hoja_biblioteca,
    leer_hoja_plantilla,
    normalizar_texto,
    parsear_archivo_biblioteca,
    parsear_archivo_plantillas,
)
from importaciones.services import (
    ImportacionInvalida,
    construir_ejemplo_plantillas,
    confirmar_importacion_biblioteca,
    confirmar_importacion_plantillas,
    previsualizar_importacion_biblioteca,
    previsualizar_importacion_plantillas,
)
from rutinas.models import RutinaPlantilla, RutinaPlantillaItem
from tenants.models import Gimnasio, Perfil


def _archivo_xlsx(wb):
    """Serializa un Workbook de openpyxl a un SimpleUploadedFile, como
    llegaría desde un <input type=file>."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile(
        "plan.xlsx", buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class ImportacionModeloTests(TestCase):
    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(
            nombre="Gimnasio Central", slug="gimnasio-central"
        )

    def test_creacion_con_defaults(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio,
            tipo=Importacion.Tipo.PLANTILLAS,
            resultado={"hojas": []},
        )
        self.assertEqual(importacion.estado, Importacion.Estado.EN_REVISION)
        self.assertIsNone(importacion.confirmado_en)
        self.assertIsNone(importacion.creado_por)

    def test_str(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio,
            tipo=Importacion.Tipo.BIBLIOTECA,
            resultado={"items": []},
        )
        self.assertIn("Gimnasio Central", str(importacion))


class ImportacionAislamientoTenantTests(TestCase):
    def test_for_gimnasio_no_mezcla_gimnasios(self):
        gimnasio_a = Gimnasio.objects.create(nombre="Gym A", slug="gym-a")
        gimnasio_b = Gimnasio.objects.create(nombre="Gym B", slug="gym-b")
        Importacion.objects.create(
            gimnasio=gimnasio_a, tipo=Importacion.Tipo.PLANTILLAS, resultado={}
        )
        Importacion.objects.create(
            gimnasio=gimnasio_b, tipo=Importacion.Tipo.PLANTILLAS, resultado={}
        )
        self.assertEqual(Importacion.objects.for_gimnasio(gimnasio_a).count(), 1)
        self.assertEqual(Importacion.objects.for_gimnasio(gimnasio_b).count(), 1)


class NormalizarTextoTests(SimpleTestCase):
    def test_lowercase_y_sin_tildes(self):
        self.assertEqual(normalizar_texto("Press de Banca"), "press de banca")
        self.assertEqual(normalizar_texto("PRÉSS DÉ BÁNCA"), "press de banca")

    def test_colapsa_espacios(self):
        self.assertEqual(normalizar_texto("  Sentadilla   libre  "), "sentadilla libre")

    def test_texto_vacio(self):
        self.assertEqual(normalizar_texto(""), "")
        self.assertEqual(normalizar_texto(None), "")


class DetectarColumnasTests(SimpleTestCase):
    def test_matchea_por_alias_case_y_acento_insensible(self):
        encabezados = ["Semana", "Día", "Ejercicio", "Series", "Repeticiones", "Descanso", "Notas"]
        campos, advertencias = detectar_columnas(encabezados, ALIAS_PLANTILLA)
        self.assertEqual(
            campos,
            {"semana": 0, "dia": 1, "ejercicio": 2, "series": 3,
             "repeticiones": 4, "descanso": 5, "notas": 6},
        )
        self.assertEqual(advertencias, [])

    def test_orden_de_columnas_no_importa(self):
        encabezados = ["Ejercicio", "Series", "Dia"]
        campos, _ = detectar_columnas(encabezados, ALIAS_PLANTILLA)
        self.assertEqual(campos, {"ejercicio": 0, "series": 1, "dia": 2})

    def test_columna_no_encontrada_no_esta_en_el_resultado(self):
        encabezados = ["Ejercicio", "Series"]
        campos, _ = detectar_columnas(encabezados, ALIAS_PLANTILLA)
        self.assertNotIn("descanso", campos)
        self.assertNotIn("semana", campos)

    def test_columna_duplicada_usa_la_primera_y_avisa(self):
        encabezados = ["Series", "Ejercicio", "Sets"]  # "Series" y "Sets" son alias de "series"
        campos, advertencias = detectar_columnas(encabezados, ALIAS_PLANTILLA)
        self.assertEqual(campos["series"], 0)
        self.assertEqual(len(advertencias), 1)
        self.assertIn("series", advertencias[0].lower())

    def test_encabezado_none_se_ignora(self):
        # Celda de header vacía (columna sin usar en la planilla real).
        encabezados = ["Ejercicio", None, "Series"]
        campos, _ = detectar_columnas(encabezados, ALIAS_PLANTILLA)
        self.assertEqual(campos, {"ejercicio": 0, "series": 2})

    def test_alias_biblioteca(self):
        encabezados = ["Nombre", "Grupo Muscular", "Video"]
        campos, _ = detectar_columnas(encabezados, ALIAS_BIBLIOTECA)
        self.assertEqual(campos, {"nombre": 0, "grupo_muscular": 1, "url_video": 2})

    def test_alias_biblioteca_acepta_categoria(self):
        """Encabezado real del primer cliente: su Excel de 748 ejercicios
        decía CATEGORÍA, que no estaba en la lista de alias, así que la
        columna no se detectaba y los 748 salían sin clasificar. Es el
        defecto que originó toda la feature de categorías por gimnasio."""
        encabezados = ["NOMBRE", "LINK", "CATEGORÍA"]
        campos, _ = detectar_columnas(encabezados, ALIAS_BIBLIOTECA)
        self.assertEqual(campos, {"nombre": 0, "url_video": 1, "grupo_muscular": 2})

    def test_alias_biblioteca_acepta_variantes_de_categoria(self):
        for encabezado in ["Categoria", "categorías", "CATEGORIAS", "Grupo"]:
            with self.subTest(encabezado=encabezado):
                campos, _ = detectar_columnas(
                    ["Nombre", encabezado], ALIAS_BIBLIOTECA
                )
                self.assertEqual(campos.get("grupo_muscular"), 1)


def _hoja_plantilla_basica():
    """Workbook en memoria con encabezados + 2 filas válidas, sin celdas
    combinadas. Reutilizado por varios tests de este módulo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Semana", "Dia", "Ejercicio", "Series", "Repeticiones", "Descanso", "Notas"])
    ws.append([1, 1, "Press de banca", 4, "8-12", "90s", ""])
    ws.append([1, 1, "Sentadilla", 3, "10", "60s", "Cuidar la técnica"])
    return ws


class LeerHojaPlantillaTests(SimpleTestCase):
    def test_lee_filas_validas(self):
        hoja = leer_hoja_plantilla(_hoja_plantilla_basica())
        self.assertEqual(len(hoja.items), 2)
        self.assertEqual(hoja.items[0], ItemParseado(
            semana=1, dia=1, orden=1, ejercicio_original="Press de banca",
            series=4, repeticiones="8-12", kilos="", descanso="90s", notas="",
            fila_excel=2,
        ))
        self.assertEqual(hoja.items[1].orden, 2)  # segundo item del mismo (semana, dia)
        self.assertEqual(hoja.filas_invalidas, [])

    def test_lee_columna_carga_como_kilos(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones", "Carga"])
        ws.append([1, "Sentadilla", 4, "8-12", "20kg"])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items[0].kilos, "20kg")

    def test_sin_columna_carga_kilos_queda_vacio(self):
        """La columna es opcional -- una hoja sin ella sigue importando
        igual que antes, solo que `kilos` queda "" ."""
        hoja = leer_hoja_plantilla(_hoja_plantilla_basica())
        self.assertEqual(hoja.items[0].kilos, "")

    def test_dias_por_semana_es_el_maximo_dia_de_filas_validas(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", 4, "8-12"])
        ws.append([3, "Sentadilla", 3, "10"])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.dias_por_semana, 3)

    def test_fila_sin_semana_cae_en_semana_1(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])  # sin columna Semana
        ws.append([1, "Press de banca", 4, "8-12"])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items[0].semana, 1)

    def test_fila_con_series_invalida_se_saltea_con_motivo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", "cuatro", "8-12"])  # "series" no numérico
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items, [])
        self.assertEqual(len(hoja.filas_invalidas), 1)
        self.assertEqual(hoja.filas_invalidas[0].fila_excel, 2)  # fila 1 = header
        self.assertIn("series", hoja.filas_invalidas[0].motivo.lower())

    def test_fila_totalmente_vacia_se_saltea_sin_motivo_ruidoso(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", 4, "8-12"])
        ws.append([None, None, None, None])
        ws.append([1, "Sentadilla", 3, "10"])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(len(hoja.items), 2)
        self.assertEqual(hoja.filas_invalidas, [])  # vacía != inválida, se ignora en silencio

    def test_columna_ejercicio_ausente_devuelve_hoja_sin_items(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Dia", "Series", "Repeticiones"])  # sin "Ejercicio"
        ws.append([1, 4, "8-12"])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items, [])
        self.assertEqual(hoja.dias_por_semana, 0)

    def test_orden_secuencial_dentro_de_semana_y_dia(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Semana", "Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, 1, "Press de banca", 4, "8-12"])
        ws.append([1, 1, "Sentadilla", 3, "10"])
        ws.append([2, 1, "Peso muerto", 3, "8"])  # otra semana: orden vuelve a 1
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual([i.orden for i in hoja.items], [1, 2, 1])

    def test_celda_combinada_de_semana_se_resuelve_por_el_ancla(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Semana", "Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, 1, "Press de banca", 4, "8-12"])
        ws.append([None, 2, "Sentadilla", 3, "10"])  # "Semana" mergeada con la fila de arriba
        ws.merge_cells("A2:A3")
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items[1].semana, 1)


class LeerHojaBibliotecaTests(SimpleTestCase):
    def test_lee_ejercicios_validos(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular", "Video"])
        ws.append(["Press de banca", "Pecho", "https://youtube.com/x"])
        ws.append(["Sentadilla", "Piernas", ""])
        items, invalidas, advertencias = leer_hoja_biblioteca(ws)
        self.assertEqual(len(items), 2)
        self.assertEqual(items[0]["nombre_original"], "Press de banca")
        self.assertEqual(items[0]["grupo_muscular_original"], "Pecho")
        self.assertEqual(items[0]["url_video"], "https://youtube.com/x")
        self.assertEqual(invalidas, [])
        self.assertEqual(advertencias, [])

    def test_fila_sin_nombre_se_saltea_con_motivo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["", "Pecho"])
        items, invalidas, _ = leer_hoja_biblioteca(ws)
        self.assertEqual(items, [])
        self.assertEqual(len(invalidas), 1)

    def test_columna_grupo_muscular_es_opcional(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])
        ws.append(["Press de banca"])
        items, _, _ = leer_hoja_biblioteca(ws)
        self.assertIsNone(items[0]["grupo_muscular_original"])


class ParsearArchivoPlantillasTests(SimpleTestCase):
    def test_multi_hoja_produce_una_hojaparseada_por_hoja(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Hombres"
        ws1.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws1.append([1, "Press de banca", 4, "8-12"])
        ws2 = wb.create_sheet("Mujeres")
        ws2.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws2.append([1, "Sentadilla", 3, "10"])

        hojas = parsear_archivo_plantillas(_archivo_xlsx(wb))

        self.assertEqual(len(hojas), 2)
        self.assertEqual({h.nombre_hoja for h in hojas}, {"Hombres", "Mujeres"})


class ParsearArchivoBibliotecaTests(SimpleTestCase):
    def test_usa_la_primera_hoja(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "Pecho"])
        items, invalidas, advertencias = parsear_archivo_biblioteca(_archivo_xlsx(wb))
        self.assertEqual(len(items), 1)
        self.assertEqual(invalidas, [])
        self.assertEqual(advertencias, [])


class ResolverNombreTests(SimpleTestCase):
    def setUp(self):
        # Índice armado a mano -- resolver_nombre es pura, no toca DB.
        self.indice = {"press de banca": "PRESS_ID", "sentadilla": "SENTADILLA_ID"}

    def test_match_exacto_tras_normalizar(self):
        resultado = resolver_nombre("press de banca", self.indice)
        self.assertEqual(resultado.tipo, "exacto")
        self.assertEqual(resultado.ejercicio, "PRESS_ID")

    def test_typo_da_ambiguo_con_candidato(self):
        resultado = resolver_nombre("sentadila", self.indice)  # falta una "l"
        self.assertEqual(resultado.tipo, "ambiguo")
        self.assertEqual(resultado.candidato, "SENTADILLA_ID")
        self.assertGreaterEqual(resultado.score, 60)

    def test_nombre_sin_relacion_da_nuevo(self):
        resultado = resolver_nombre("hip thrust", self.indice)
        self.assertEqual(resultado.tipo, "nuevo")
        self.assertIsNone(resultado.candidato)

    def test_indice_vacio_siempre_da_nuevo(self):
        resultado = resolver_nombre("cualquier cosa", {})
        self.assertEqual(resultado.tipo, "nuevo")


# `ResolverGrupoMuscularTests` se retiró el 2026-08-26 junto con
# `resolver_grupo_muscular`: matcheaba texto contra un `TextChoices` global de
# 8 valores más un diccionario de alias fijo. Con el catálogo de categorías
# por gimnasio no hay lista global contra la cual matchear. Lo reemplaza
# `ResolverCategoriasTests`, que además cubre el dedupe difuso.


class ConstruirIndiceEjerciciosTests(TestCase):
    def test_indexa_por_nombre_normalizado_y_aisla_por_tenant(self):
        gimnasio_a = Gimnasio.objects.create(nombre="Gym A", slug="gym-a")
        gimnasio_b = Gimnasio.objects.create(nombre="Gym B", slug="gym-b")
        ejercicio_a = Ejercicio.objects.create(
            gimnasio=gimnasio_a, nombre="Press de Banca",
        )
        Ejercicio.objects.create(
            gimnasio=gimnasio_b, nombre="Sentadilla",
        )
        indice = construir_indice_ejercicios(gimnasio_a)
        self.assertEqual(indice, {"press de banca": ejercicio_a})


class PrevisualizarImportacionPlantillasTests(TestCase):
    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user(username="staff", password="clave12345")
        self.ejercicio_existente = Ejercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Sentadilla",
        )

    def _archivo_dos_hojas(self):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Hombres"
        ws1.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws1.append([1, "Press de banca", 4, "8-12"])
        ws1.append([1, "sentadila", 3, "10"])  # typo de un ejercicio ya cargado
        ws2 = wb.create_sheet("Mujeres")
        ws2.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws2.append([1, "Press de banca", 3, "10-12"])  # mismo ejercicio nuevo, otra hoja
        return _archivo_xlsx(wb)

    def test_crea_importacion_en_revision_sin_tocar_rutinaplantilla(self):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=self._archivo_dos_hojas(), usuario=self.usuario,
        )
        self.assertEqual(importacion.tipo, Importacion.Tipo.PLANTILLAS)
        self.assertEqual(importacion.estado, Importacion.Estado.EN_REVISION)
        self.assertEqual(importacion.gimnasio, self.gimnasio)
        self.assertEqual(importacion.creado_por, self.usuario)
        self.assertEqual(RutinaPlantilla.objects.count(), 0)

    def test_resultado_tiene_una_entrada_por_hoja(self):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=self._archivo_dos_hojas(), usuario=self.usuario,
        )
        nombres_hoja = {h["nombre_hoja"] for h in importacion.resultado["hojas"]}
        self.assertEqual(nombres_hoja, {"Hombres", "Mujeres"})

    def test_ejercicio_repetido_entre_hojas_se_resuelve_una_sola_vez(self):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=self._archivo_dos_hojas(), usuario=self.usuario,
        )
        # "Press de banca" aparece en las dos hojas -> una sola entrada en ejercicios_distintos
        self.assertIn("press de banca", importacion.resultado["ejercicios_distintos"])
        entrada = importacion.resultado["ejercicios_distintos"]["press de banca"]
        self.assertEqual(entrada["tipo"], "nuevo")

    def test_typo_de_ejercicio_existente_da_ambiguo_con_candidato(self):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=self._archivo_dos_hojas(), usuario=self.usuario,
        )
        entrada = importacion.resultado["ejercicios_distintos"]["sentadila"]
        self.assertEqual(entrada["tipo"], "ambiguo")
        self.assertEqual(entrada["candidato_id"], self.ejercicio_existente.pk)

    def test_archivo_no_valido_lanza_importacioninvalida(self):
        archivo_roto = SimpleUploadedFile("plan.xlsx", b"esto no es un xlsx")
        with self.assertRaises(ImportacionInvalida):
            previsualizar_importacion_plantillas(
                gimnasio=self.gimnasio, archivo=archivo_roto, usuario=self.usuario,
            )


class ConfirmarImportacionPlantillasTests(TestCase):
    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.usuario = User.objects.create_user(username="staff", password="clave12345")
        self.ejercicio_existente = Ejercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Sentadilla",
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", 4, "8-12"])
        ws.append([1, "sentadila", 3, "10"])
        self.importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def _decisiones_completas(self, accion_sentadila="usar_existente"):
        return {
            "hojas": [{"nombre_hoja": "Hombres", "incluir": True, "objetivo": "Hipertrofia", "nivel": "principiante"}],
            "ejercicios": {
                "press de banca": {"accion": "crear_nuevo", "categoria_id": self.pecho.pk},
                "sentadila": {
                    "accion": accion_sentadila,
                    "ejercicio_id": self.ejercicio_existente.pk if accion_sentadila == "usar_existente" else None,
                    "categoria_id": self.piernas.pk if accion_sentadila == "crear_nuevo" else None,
                },
            },
        }

    def test_crea_una_plantilla_por_hoja_incluida(self):
        plantillas = confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio,
            decisiones=self._decisiones_completas(),
        )
        self.assertEqual(len(plantillas), 1)
        self.assertEqual(plantillas[0].nombre, "Hombres")
        self.assertEqual(plantillas[0].objetivo, "Hipertrofia")
        self.assertEqual(plantillas[0].nivel, "principiante")
        self.assertEqual(plantillas[0].dias_por_semana, 1)
        self.assertEqual(plantillas[0].items.count(), 2)

    def test_usar_existente_no_duplica_el_ejercicio(self):
        confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio,
            decisiones=self._decisiones_completas(accion_sentadila="usar_existente"),
        )
        self.assertEqual(Ejercicio.objects.filter(nombre__iexact="sentadilla").count(), 1)
        item_sentadilla = RutinaPlantillaItem.objects.get(ejercicio=self.ejercicio_existente)
        self.assertEqual(item_sentadilla.series, 3)

    def test_crear_nuevo_crea_exactamente_un_ejercicio(self):
        confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio,
            decisiones=self._decisiones_completas(),
        )
        self.assertEqual(
            Ejercicio.objects.filter(gimnasio=self.gimnasio, nombre="Press de banca").count(), 1
        )

    def test_marca_la_importacion_como_confirmada(self):
        confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio,
            decisiones=self._decisiones_completas(),
        )
        self.importacion.refresh_from_db()
        self.assertEqual(self.importacion.estado, Importacion.Estado.CONFIRMADA)
        self.assertIsNotNone(self.importacion.confirmado_en)

    def test_confirmar_dos_veces_falla_sin_duplicar(self):
        confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio,
            decisiones=self._decisiones_completas(),
        )
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=self.gimnasio,
                decisiones=self._decisiones_completas(),
            )
        self.assertEqual(RutinaPlantilla.objects.count(), 1)

    def test_importacion_de_otro_gimnasio_falla(self):
        otro_gimnasio = Gimnasio.objects.create(nombre="Otro", slug="otro")
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=otro_gimnasio,
                decisiones=self._decisiones_completas(),
            )

    def test_hoja_no_incluida_no_crea_plantilla(self):
        decisiones = self._decisiones_completas()
        decisiones["hojas"][0]["incluir"] = False
        plantillas = confirmar_importacion_plantillas(
            importacion=self.importacion, gimnasio=self.gimnasio, decisiones=decisiones,
        )
        self.assertEqual(plantillas, [])
        self.assertEqual(RutinaPlantilla.objects.count(), 0)

    def test_decisiones_hojas_incompletas_falla(self):
        # Simula un form de confirmación (Tarea 9) donde faltó la decisión
        # de una hoja -- p. ej. un checkbox sin marcar que no llegó en el
        # POST. No debe saltearse en silencio.
        decisiones = self._decisiones_completas()
        decisiones["hojas"] = []
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )
        self.assertEqual(RutinaPlantilla.objects.count(), 0)
        self.importacion.refresh_from_db()
        self.assertEqual(self.importacion.estado, Importacion.Estado.EN_REVISION)

    def test_categoria_de_otro_gimnasio_o_inexistente_falla(self):
        decisiones = self._decisiones_completas()
        decisiones["ejercicios"]["press de banca"]["categoria_id"] = 999999
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )

    def test_nivel_invalido_falla(self):
        decisiones = self._decisiones_completas()
        decisiones["hojas"][0]["nivel"] = "experto-supremo"
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )

    def test_usar_existente_con_ejercicio_de_otro_gimnasio_falla(self):
        otro_gimnasio = Gimnasio.objects.create(nombre="Otro", slug="otro")
        ejercicio_de_otro_gimnasio = Ejercicio.objects.create(
            gimnasio=otro_gimnasio, nombre="Sentadilla",
        )
        decisiones = self._decisiones_completas(accion_sentadila="usar_existente")
        decisiones["ejercicios"]["sentadila"]["ejercicio_id"] = ejercicio_de_otro_gimnasio.pk
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=self.importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )
        self.assertEqual(RutinaPlantilla.objects.count(), 0)

    def test_mismo_ejercicio_nuevo_en_dos_hojas_crea_uno_solo(self):
        # A diferencia del fixture de setUp (una sola hoja), acá el mismo
        # nombre aparece en DOS hojas distintas -- el memo de
        # `_obtener_ejercicio` tiene que estar scopeado a todo el confirm,
        # no por hoja.
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Hombres"
        ws1.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws1.append([1, "Peso muerto", 4, "8-12"])
        ws2 = wb.create_sheet("Mujeres")
        ws2.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws2.append([1, "Peso muerto", 3, "10"])
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        decisiones = {
            "hojas": [
                {"nombre_hoja": "Hombres", "incluir": True, "objetivo": "Fuerza", "nivel": "principiante"},
                {"nombre_hoja": "Mujeres", "incluir": True, "objetivo": "Fuerza", "nivel": "principiante"},
            ],
            "ejercicios": {
                "peso muerto": {"accion": "crear_nuevo", "categoria_id": self.piernas.pk},
            },
        }
        plantillas = confirmar_importacion_plantillas(
            importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones,
        )
        self.assertEqual(
            Ejercicio.objects.filter(gimnasio=self.gimnasio, nombre="Peso muerto").count(), 1
        )
        ejercicio = Ejercicio.objects.get(gimnasio=self.gimnasio, nombre="Peso muerto")
        items = RutinaPlantillaItem.objects.filter(ejercicio=ejercicio)
        self.assertEqual(items.count(), 2)
        self.assertEqual({p.pk for p in plantillas}, set(items.values_list("rutina_id", flat=True)))

    def test_falla_a_mitad_de_transaccion_no_deja_datos_parciales(self):
        # Dos hojas: la primera es válida y alcanzaría a crear su plantilla
        # e ítems antes de que la segunda dispare la validación de grupo
        # muscular. Todo el atomic() tiene que revertirse, no solo la
        # segunda hoja.
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Hombres"
        ws1.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws1.append([1, "Press de banca", 4, "8-12"])
        ws2 = wb.create_sheet("Mujeres")
        ws2.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws2.append([1, "Peso muerto", 3, "8"])
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        ejercicios_antes = Ejercicio.objects.count()
        decisiones = {
            "hojas": [
                {"nombre_hoja": "Hombres", "incluir": True, "objetivo": "Hipertrofia", "nivel": "principiante"},
                {"nombre_hoja": "Mujeres", "incluir": True, "objetivo": "Fuerza", "nivel": "principiante"},
            ],
            "ejercicios": {
                "press de banca": {"accion": "crear_nuevo", "categoria_id": self.pecho.pk},
                "peso muerto": {"accion": "crear_nuevo", "categoria_id": 999999},
            },
        }
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )
        self.assertEqual(RutinaPlantilla.objects.count(), 0)
        self.assertEqual(RutinaPlantillaItem.objects.count(), 0)
        self.assertEqual(Ejercicio.objects.count(), ejercicios_antes)
        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.EN_REVISION)


class ConfirmarImportacionPlantillasConCargaTests(TestCase):
    """La columna 'Carga' del Excel llega hasta `RutinaPlantillaItem.kilos`
    de punta a punta (parsing -> JSON del preview -> confirm)."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.usuario = User.objects.create_user(username="staff", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones", "Carga"])
        ws.append([1, "Sentadilla", 4, "8-12", "20kg"])
        self.importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def test_kilos_llega_al_item_creado(self):
        plantillas = confirmar_importacion_plantillas(
            importacion=self.importacion,
            gimnasio=self.gimnasio,
            decisiones={
                "hojas": [{"nombre_hoja": "Hombres", "incluir": True, "objetivo": "Hipertrofia", "nivel": "principiante"}],
                "ejercicios": {
                    "sentadilla": {"accion": "crear_nuevo", "categoria_id": self.piernas.pk},
                },
            },
        )
        item = plantillas[0].items.get()
        self.assertEqual(item.kilos, "20kg")


class ImportacionBibliotecaTests(TestCase):
    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.usuario = User.objects.create_user(username="staff", password="clave12345")
        self.ejercicio_existente = Ejercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Sentadilla",
        )

    def _archivo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular", "Video"])
        ws.append(["Press de banca", "Pecho", "https://youtube.com/x"])
        ws.append(["sentadila", "Piernas", ""])  # typo de la ya existente
        return _archivo_xlsx(wb)

    def test_previsualizar_no_crea_ejercicios(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        self.assertEqual(importacion.tipo, Importacion.Tipo.BIBLIOTECA)
        self.assertEqual(Ejercicio.objects.count(), 1)  # solo la que ya existía

    def test_previsualizar_resuelve_la_categoria_automaticamente(self):
        """La columna del archivo dice "Pecho" y el gimnasio ya tiene esa
        categoría: se reusa, no se crea una segunda."""
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        item = next(i for i in importacion.resultado["items"] if i["nombre_original"] == "Press de banca")
        self.assertEqual(item["categoria_resuelta"]["tipo"], "existente")
        self.assertEqual(item["categoria_resuelta"]["categoria_id"], self.pecho.pk)

    def test_confirmar_crea_solo_los_ejercicios_nuevos(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        creados = confirmar_importacion_biblioteca(
            importacion=importacion, gimnasio=self.gimnasio,
            decisiones={"items": {
                "press de banca": {"incluir": True, "categoria_id": self.pecho.pk},
                "sentadila": {"incluir": False, "categoria_id": None},
            }},
        )
        self.assertEqual(len(creados), 1)
        self.assertEqual(Ejercicio.objects.filter(gimnasio=self.gimnasio).count(), 2)

    def test_confirmar_dos_veces_falla(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        decisiones = {"items": {
            "press de banca": {"incluir": True, "categoria_id": self.pecho.pk},
            "sentadila": {"incluir": False, "categoria_id": None},
        }}
        confirmar_importacion_biblioteca(importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones)
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_biblioteca(importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones)

    def test_confirmar_grupo_muscular_invalido_no_crea_nada(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        ejercicios_antes = Ejercicio.objects.count()
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_biblioteca(
                importacion=importacion, gimnasio=self.gimnasio,
                decisiones={"items": {
                    "press de banca": {"incluir": True, "categoria_id": 999999},
                    "sentadila": {"incluir": False, "categoria_id": None},
                }},
            )
        self.assertEqual(Ejercicio.objects.count(), ejercicios_antes)

    def test_confirmar_con_gimnasio_distinto_falla(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        otro_gimnasio = Gimnasio.objects.create(nombre="Otro Gym", slug="otro-gym")
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_biblioteca(
                importacion=importacion, gimnasio=otro_gimnasio,
                decisiones={"items": {
                    "press de banca": {"incluir": True, "categoria_id": self.pecho.pk},
                    "sentadila": {"incluir": False, "categoria_id": None},
                }},
            )

    def test_confirmar_decision_faltante_da_error_claro(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_biblioteca(
                importacion=importacion, gimnasio=self.gimnasio,
                decisiones={"items": {
                    "press de banca": {"incluir": True, "categoria_id": self.pecho.pk},
                    # falta la decisión de "sentadila"
                }},
            )

    def test_confirmar_match_ambiguo_crea_nuevo_ejercicio(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        item = next(i for i in importacion.resultado["items"] if i["nombre_original"] == "sentadila")
        self.assertEqual(item["match"]["tipo"], "ambiguo")
        creados = confirmar_importacion_biblioteca(
            importacion=importacion, gimnasio=self.gimnasio,
            decisiones={"items": {
                "press de banca": {"incluir": False, "categoria_id": None},
                "sentadila": {"incluir": True, "categoria_id": self.piernas.pk},
            }},
        )
        self.assertEqual(len(creados), 1)
        self.assertEqual(
            Ejercicio.objects.filter(gimnasio=self.gimnasio, nombre="sentadila").count(), 1,
        )


class SubirPlantillasFormTests(TestCase):
    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")

    def test_acepta_xlsx(self):
        wb = openpyxl.Workbook()
        archivo = _archivo_xlsx(wb)
        from importaciones.forms import SubirPlantillasForm
        form = SubirPlantillasForm(data={}, files={"archivo": archivo}, gimnasio=self.gimnasio)
        self.assertTrue(form.is_valid(), form.errors)

    def test_rechaza_extension_invalida(self):
        archivo = SimpleUploadedFile("plan.csv", b"a,b,c", content_type="text/csv")
        from importaciones.forms import SubirPlantillasForm
        form = SubirPlantillasForm(data={}, files={"archivo": archivo}, gimnasio=self.gimnasio)
        self.assertFalse(form.is_valid())
        self.assertIn("archivo", form.errors)

    def test_biblioteca_tambien_valida_extension(self):
        archivo = SimpleUploadedFile("plan.txt", b"nada")
        from importaciones.forms import SubirBibliotecaForm
        form = SubirBibliotecaForm(data={}, files={"archivo": archivo}, gimnasio=self.gimnasio)
        self.assertFalse(form.is_valid())


class HojaMetadataFormSetTests(SimpleTestCase):
    def test_requiere_objetivo_y_nivel(self):
        from importaciones.forms import HojaMetadataFormSet
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Hombres", "form-0-incluir": "on",
            "form-0-objetivo": "", "form-0-nivel": "",
        }
        formset = HojaMetadataFormSet(datos)
        self.assertFalse(formset.is_valid())

    def test_valido_con_todos_los_campos(self):
        from importaciones.forms import HojaMetadataFormSet
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Hombres", "form-0-incluir": "on",
            "form-0-objetivo": "Hipertrofia", "form-0-nivel": "principiante",
        }
        formset = HojaMetadataFormSet(datos)
        self.assertTrue(formset.is_valid(), formset.errors)


class ResolucionEjercicioFormSetTests(SimpleTestCase):
    def test_crear_nuevo_requiere_grupo_muscular(self):
        from importaciones.forms import ResolucionEjercicioFormSet
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_normalizado": "hip thrust", "form-0-accion": "crear_nuevo",
            "form-0-grupo_muscular": "",
        }
        formset = ResolucionEjercicioFormSet(datos)
        self.assertFalse(formset.is_valid())

    def test_usar_existente_no_requiere_grupo_muscular(self):
        from importaciones.forms import ResolucionEjercicioFormSet
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_normalizado": "sentadila", "form-0-accion": "usar_existente",
            "form-0-ejercicio_existente_id": "7", "form-0-grupo_muscular": "",
        }
        formset = ResolucionEjercicioFormSet(datos)
        self.assertTrue(formset.is_valid(), formset.errors)


class ImportacionPlantillasViewsTests(TestCase):
    def setUp(self):
        self.gimnasio_a = Gimnasio.objects.create(nombre="Gym A", slug="gym-a")
        self.gimnasio_b = Gimnasio.objects.create(nombre="Gym B", slug="gym-b")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio_a, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio_a, nombre="Piernas"
        )

        self.staff_a = User.objects.create_user(username="staff_a", password="clave12345")
        Perfil.objects.create(usuario=self.staff_a, gimnasio=self.gimnasio_a, rol=Perfil.Rol.STAFF)

        self.staff_b = User.objects.create_user(username="staff_b", password="clave12345")
        Perfil.objects.create(usuario=self.staff_b, gimnasio=self.gimnasio_b, rol=Perfil.Rol.STAFF)

        self.usuario_alumno = User.objects.create_user(username="usuario_alumno", password="clave12345")
        Perfil.objects.create(usuario=self.usuario_alumno, gimnasio=self.gimnasio_a, rol=Perfil.Rol.ALUMNO)

    def _archivo_valido(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", 4, "8-12"])
        return _archivo_xlsx(wb)

    def test_anonimo_redirige_a_login(self):
        response = self.client.get(reverse("importaciones:plantillas_subir"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_alumno_recibe_403(self):
        self.client.login(username="usuario_alumno", password="clave12345")
        response = self.client.get(reverse("importaciones:plantillas_subir"))
        self.assertEqual(response.status_code, 403)

    def test_subir_archivo_invalido_no_crea_importacion(self):
        self.client.login(username="staff_a", password="clave12345")
        archivo = SimpleUploadedFile("plan.txt", b"nada")
        response = self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": archivo},
        )
        self.assertEqual(response.status_code, 200)  # re-renderiza con error
        self.assertEqual(Importacion.objects.count(), 0)

    def test_preview_de_otro_gimnasio_da_404(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_b, tipo=Importacion.Tipo.PLANTILLAS, resultado={"hojas": [], "ejercicios_distintos": {}},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_preview_de_importacion_ya_confirmada_da_404(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_a, tipo=Importacion.Tipo.PLANTILLAS,
            estado=Importacion.Estado.CONFIRMADA, resultado={"hojas": [], "ejercicios_distintos": {}},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_preview_muestra_el_video_de_cada_ejercicio(self):
        """El archivo real del cliente trae una columna LINK con el video de
        cada ejercicio, se parsea bien, pero el preview no la mostraba en
        ningún lado -- no había forma de detectar un link mal pegado antes de
        confirmar (reporte del 2026-08-27)."""
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["NOMBRE", "LINK", "CATEGORÍA"])
        ws.append(["Press de banca", "https://youtu.be/abc123", "Pecho"])
        ws.append(["Zancada", "", "Piernas"])
        self.client.post(reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)})
        importacion = Importacion.objects.get()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "https://youtu.be/abc123")
        self.assertContains(response, "Sin video")

    def test_flujo_completo_subir_preview_confirmar(self):
        self.client.login(username="staff_a", password="clave12345")

        response = self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": self._archivo_valido()},
        )
        self.assertEqual(response.status_code, 302)
        importacion = Importacion.objects.get()
        self.assertRedirects(
            response, reverse("importaciones:plantillas_hojas", args=[importacion.pk])
        )

        # Paso nuevo: elegir qué hojas del archivo son planes.
        response = self.client.get(response.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hombres")

        response = self.client.post(
            reverse("importaciones:plantillas_hojas", args=[importacion.pk]),
            {"hojas": ["Hombres"]},
        )
        self.assertRedirects(
            response, reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )

        response = self.client.get(response.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Hombres")
        self.assertContains(response, "Press de banca")

        datos_confirmacion = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Hombres", "form-0-incluir": "on",
            "form-0-objetivo": "Hipertrofia", "form-0-nivel": "principiante",
            "ejercicios-TOTAL_FORMS": "1", "ejercicios-INITIAL_FORMS": "1",
            "ejercicios-0-nombre_normalizado": "press de banca",
            "ejercicios-0-accion": "crear_nuevo",
            "ejercicios-0-categoria": self.pecho.pk,
        }
        response = self.client.post(
            reverse("importaciones:plantillas_preview", args=[importacion.pk]), datos_confirmacion,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(RutinaPlantilla.objects.count(), 1)
        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.CONFIRMADA)

        # Reabrir el preview de una importación ya confirmada da 404.
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_descartar_marca_la_importacion_como_descartada(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_a, tipo=Importacion.Tipo.PLANTILLAS, resultado={"hojas": [], "ejercicios_distintos": {}},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.post(
            reverse("importaciones:plantillas_descartar", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 302)
        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.DESCARTADA)


class ImportacionBibliotecaViewsTests(TestCase):
    def setUp(self):
        self.gimnasio_a = Gimnasio.objects.create(nombre="Gym A", slug="gym-a")
        self.gimnasio_b = Gimnasio.objects.create(nombre="Gym B", slug="gym-b")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio_a, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio_a, nombre="Piernas"
        )

        self.staff_a = User.objects.create_user(username="staff_a", password="clave12345")
        Perfil.objects.create(usuario=self.staff_a, gimnasio=self.gimnasio_a, rol=Perfil.Rol.STAFF)

        self.usuario_alumno = User.objects.create_user(username="usuario_alumno", password="clave12345")
        Perfil.objects.create(usuario=self.usuario_alumno, gimnasio=self.gimnasio_a, rol=Perfil.Rol.ALUMNO)

    def _archivo_valido(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "Pecho"])
        return _archivo_xlsx(wb)

    def test_anonimo_redirige_a_login(self):
        response = self.client.get(reverse("importaciones:biblioteca_subir"))
        self.assertEqual(response.status_code, 302)
        self.assertIn("/accounts/login/", response.url)

    def test_alumno_recibe_403(self):
        self.client.login(username="usuario_alumno", password="clave12345")
        response = self.client.get(reverse("importaciones:biblioteca_subir"))
        self.assertEqual(response.status_code, 403)

    def test_subir_archivo_invalido_no_crea_importacion(self):
        self.client.login(username="staff_a", password="clave12345")
        archivo = SimpleUploadedFile("ejercicios.txt", b"nada")
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": archivo},
        )
        self.assertEqual(response.status_code, 200)  # re-renderiza con error
        self.assertEqual(Importacion.objects.count(), 0)

    def test_preview_de_otro_gimnasio_da_404(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_b, tipo=Importacion.Tipo.BIBLIOTECA, resultado={"items": []},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_preview_de_importacion_ya_confirmada_da_404(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_a, tipo=Importacion.Tipo.BIBLIOTECA,
            estado=Importacion.Estado.CONFIRMADA, resultado={"items": []},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_flujo_completo_subir_preview_confirmar(self):
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "Pecho"])

        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        self.assertEqual(response.status_code, 302)
        importacion = Importacion.objects.get()
        self.assertRedirects(
            response, reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

        response = self.client.get(response.url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Press de banca")

        # "Press de banca" resolvió grupo_muscular automáticamente ("pecho")
        # y no necesita entrada en las resoluciones manuales.
        datos = {"resoluciones": "{}"}
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.filter(nombre="Press de banca").count(), 1)
        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.CONFIRMADA)

        # Reabrir el preview de una importación ya confirmada da 404.
        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 404)

    def test_flujo_con_resolucion_manual_de_categoria(self):
        # "hip thrust" no tiene grupo muscular en el archivo -> requiere
        # entrada en el formset de resolución manual del preview.
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])
        ws.append(["Hip thrust"])

        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "hip thrust")

        datos = {"resoluciones": json.dumps({"hip thrust": {"categoria_id": self.piernas.pk}})}
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.filter(nombre="Hip thrust").count(), 1)
        self.assertEqual(
            Ejercicio.objects.get(nombre="Hip thrust").categoria, self.piernas
        )

    def test_falta_resolver_un_pendiente_no_confirma(self):
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])
        ws.append(["Hip thrust"])
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": "{}"},  # "hip thrust" queda sin resolver
        )
        self.assertEqual(response.status_code, 200)  # re-renderiza con error
        self.assertEqual(Ejercicio.objects.count(), 0)

    def test_resoluciones_con_json_invalido_no_confirma_y_muestra_error(self):
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": self._archivo_valido()},
        )
        importacion = Importacion.objects.get()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": "not json"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Formato de resoluciones inválido.")
        self.assertEqual(Ejercicio.objects.count(), 0)

    def test_resoluciones_con_grupo_muscular_invalido_no_confirma_y_muestra_error(self):
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])
        ws.append(["Hip thrust"])
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({"hip thrust": {"categoria_id": "no_es_un_entero"}})},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Categoría inválida.")
        self.assertEqual(Ejercicio.objects.count(), 0)

    def test_preview_lista_filas_invalidas_con_motivo(self):
        # Regla global no negociable: "filas inválidas se saltean y se
        # listan con motivo" -- nunca se descartan en silencio.
        self.client.login(username="staff_a", password="clave12345")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "Pecho"])
        ws.append(["", "Pecho"])  # sin nombre -> fila inválida

        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        self.assertEqual(len(importacion.resultado["filas_invalidas"]), 1)

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "Falta el nombre del ejercicio")

    def test_descartar_marca_la_importacion_como_descartada_y_redirige_a_subir(self):
        importacion = Importacion.objects.create(
            gimnasio=self.gimnasio_a, tipo=Importacion.Tipo.BIBLIOTECA, resultado={"items": []},
        )
        self.client.login(username="staff_a", password="clave12345")
        response = self.client.post(
            reverse("importaciones:biblioteca_descartar", args=[importacion.pk])
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("importaciones:biblioteca_subir"))
        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.DESCARTADA)

    def _archivo_con_ambiguo(self, gimnasio):
        Ejercicio.objects.create(
            gimnasio=gimnasio, nombre="Sentadilla", grupo_muscular="piernas",
        )
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])
        ws.append(["Sentadila"])  # typo -> match ambiguo contra "Sentadilla"
        return _archivo_xlsx(wb)

    def test_preview_muestra_candidato_y_score_para_match_ambiguo(self):
        self.client.login(username="staff_a", password="clave12345")
        self.client.post(
            reverse("importaciones:biblioteca_subir"),
            {"archivo": self._archivo_con_ambiguo(self.gimnasio_a)},
        )
        importacion = Importacion.objects.get()
        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "Sentadilla")  # nombre del candidato
        # Score de rapidfuzz para "sentadila" vs "sentadilla" (WRatio),
        # confirmado corriendo `resolver_nombre` directamente: 94.
        self.assertContains(response, "94")

    def test_ambiguo_usar_existente_no_crea_ejercicio_nuevo(self):
        self.client.login(username="staff_a", password="clave12345")
        self.client.post(
            reverse("importaciones:biblioteca_subir"),
            {"archivo": self._archivo_con_ambiguo(self.gimnasio_a)},
        )
        importacion = Importacion.objects.get()
        datos = {
            "resoluciones": json.dumps({"sentadila": {"accion": "usar_existente"}}),
        }
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.filter(gimnasio=self.gimnasio_a).count(), 1)

    def test_ambiguo_crear_nuevo_requiere_grupo_muscular_y_crea_ejercicio(self):
        self.client.login(username="staff_a", password="clave12345")
        self.client.post(
            reverse("importaciones:biblioteca_subir"),
            {"archivo": self._archivo_con_ambiguo(self.gimnasio_a)},
        )
        importacion = Importacion.objects.get()
        datos = {
            "resoluciones": json.dumps(
                {"sentadila": {"accion": "crear_nuevo", "categoria_id": self.piernas.pk}}
            ),
        }
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.filter(gimnasio=self.gimnasio_a).count(), 2)
        self.assertTrue(
            Ejercicio.objects.filter(gimnasio=self.gimnasio_a, nombre="Sentadila").exists()
        )

    def test_ambiguo_sin_resolver_no_confirma(self):
        self.client.login(username="staff_a", password="clave12345")
        self.client.post(
            reverse("importaciones:biblioteca_subir"),
            {"archivo": self._archivo_con_ambiguo(self.gimnasio_a)},
        )
        importacion = Importacion.objects.get()
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": "{}"},
        )
        self.assertEqual(response.status_code, 200)  # re-renderiza con error
        self.assertEqual(Ejercicio.objects.filter(gimnasio=self.gimnasio_a).count(), 1)


class RegresionCamposDelPostTests(TestCase):
    """El confirm POST manda solo decisiones (objetivo/nivel por hoja +
    resolución por ejercicio distinto), nunca el dataset entero -- por
    diseño (spec §2) no debería acercarse jamás al límite default de
    Django (1000 campos por POST) para una hoja de tamaño realista, sin
    importar cuántas filas tenga.

    NOTA (Tarea 12): la planilla original de este test (500 filas, cada una
    con un nombre de ejercicio DISTINTO -- "Ejercicio {i}" para las 500)
    contradice la premisa que el propio test dice validar: con 500
    ejercicios distintos el confirm POST manda ~1500 campos (3 por
    ejercicio × 500) y SÍ supera el límite default de 1000, tirando
    `TooManyFieldsSent`. Ninguna plantilla real tiene 500 ejercicios
    completamente distintos en una sola hoja -- una planilla de ese tamaño
    normalmente repite un vocabulario acotado de ejercicios a lo largo de
    varias semanas/días (ver el ejemplo de la spec: "4 semanas × 5 días × 6
    ejercicios × 2 hojas ~240 filas"). Este test usa 500 filas que reciclan
    un pool de 20 ejercicios distintos -- volumen de filas realista, sin
    caer en el caso patológico que rompe el propio invariante que se quiere
    demostrar. Ver ISSUES.md `[2026-07-28]` para el detalle."""

    CANTIDAD_FILAS = 500
    CANTIDAD_EJERCICIOS_DISTINTOS = 20

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)

    def test_hoja_de_500_filas_no_rompe_el_confirm_post(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Full body"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        for i in range(self.CANTIDAD_FILAS):
            nombre = f"Ejercicio {i % self.CANTIDAD_EJERCICIOS_DISTINTOS}"
            ws.append([1, nombre, 3, "10"])

        self.client.login(username="staff", password="clave12345")
        response = self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        self.assertEqual(
            len(importacion.resultado["ejercicios_distintos"]),
            self.CANTIDAD_EJERCICIOS_DISTINTOS,
        )

        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Full body", "form-0-incluir": "on",
            "form-0-objetivo": "General", "form-0-nivel": "principiante",
            "ejercicios-TOTAL_FORMS": str(self.CANTIDAD_EJERCICIOS_DISTINTOS),
            "ejercicios-INITIAL_FORMS": str(self.CANTIDAD_EJERCICIOS_DISTINTOS),
        }
        for i, nombre in enumerate(importacion.resultado["ejercicios_distintos"]):
            datos[f"ejercicios-{i}-nombre_normalizado"] = nombre
            datos[f"ejercicios-{i}-accion"] = "crear_nuevo"
            datos[f"ejercicios-{i}-categoria"] = self.pecho.pk

        response = self.client.post(
            reverse("importaciones:plantillas_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            RutinaPlantilla.objects.get().items.count(), self.CANTIDAD_FILAS
        )


class RegresionCamposPostBibliotecaTests(TestCase):
    """El confirm POST de biblioteca manda las resoluciones como un único
    campo JSON -- a diferencia de plantillas (ver ISSUES.md [2026-07-28]),
    el conteo de campos del POST no escala con la cantidad de ejercicios
    pendientes de resolución manual, así que una biblioteca inicial de
    miles de ejercicios (escenario real, a diferencia de una plantilla)
    no puede romper contra DATA_UPLOAD_MAX_NUMBER_FIELDS."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)

    def test_600_ejercicios_pendientes_no_rompe_el_confirm_post(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])  # sin columna Grupo Muscular -> todos pendientes
        for i in range(600):
            ws.append([f"Ejercicio {i}"])

        self.client.login(username="staff", password="clave12345")
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        self.assertEqual(len(importacion.resultado["items"]), 600)

        resoluciones = {
            f"ejercicio {i}": {"categoria_id": self.pecho.pk} for i in range(600)
        }
        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps(resoluciones)},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.count(), 600)


class SinDefaultSilenciosoDeGrupoMuscularYNivelTests(TestCase):
    """Fix post-review, hallazgo 1: ni `grupo_muscular` (ejercicio nuevo) ni
    `nivel` (hoja) pueden quedar con una opción real pre-seleccionada en el
    <select> -- eso viola el constraint no negociable de que el staff SIEMPRE
    elige el grupo muscular de un ejercicio nuevo a mano. Antes del fix, el
    <select> no tenía ninguna opción en blanco, así que el navegador
    pre-seleccionaba (y mandaba) la primera choice real ("pecho" /
    "principiante") sin que el staff la tocara."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)
        self.client.login(username="staff", password="clave12345")

    def _importacion_con_ejercicio_nuevo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", 4, "8-12"])
        self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        return Importacion.objects.get()

    def test_preview_no_preselecciona_ninguna_opcion_real_en_los_selects(self):
        importacion = self._importacion_con_ejercicio_nuevo()
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        # La opción en blanco es la que queda "selected" -- ni "pecho" (1ra
        # choice de grupo_muscular) ni "principiante" (1ra choice de nivel).
        self.assertContains(response, '<option value="" selected>---------</option>', count=2)
        self.assertNotContains(response, '<option value="pecho" selected>')
        self.assertNotContains(response, '<option value="principiante" selected>')

    def test_confirmar_sin_elegir_grupo_muscular_no_crea_nada(self):
        # Simula un navegador real: el staff no tocó el <select>, así que
        # se manda la opción en blanco pre-seleccionada, no "pecho".
        importacion = self._importacion_con_ejercicio_nuevo()
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Hombres", "form-0-incluir": "on",
            "form-0-objetivo": "Hipertrofia", "form-0-nivel": "principiante",
            "ejercicios-TOTAL_FORMS": "1", "ejercicios-INITIAL_FORMS": "1",
            "ejercicios-0-nombre_normalizado": "press de banca",
            "ejercicios-0-accion": "crear_nuevo",
            "ejercicios-0-categoria": "",
        }
        response = self.client.post(
            reverse("importaciones:plantillas_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 200)  # re-renderiza con error, no redirige
        self.assertContains(response, "Elegí una categoría")
        self.assertEqual(RutinaPlantilla.objects.count(), 0)
        self.assertEqual(Ejercicio.objects.count(), 0)


class HojaExcluidaPorColumnaFaltanteTests(TestCase):
    """Fix post-review, hallazgo 2: una hoja sin una columna requerida
    (ejercicio/series/repeticiones) queda con 0 items -- antes, sin ningún
    motivo registrado, y con `incluir` pre-tildado por default, así que
    confirmarla creaba una `RutinaPlantilla` vacía en silencio."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)
        self.client.login(username="staff", password="clave12345")

    def _archivo_con_columna_ejercicio_mal_escrita(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercico", "Series", "Repeticiones"])  # typo: "Ejercico"
        ws.append([1, "Press de banca", 4, "8-12"])
        return _archivo_xlsx(wb)

    def test_preview_registra_y_muestra_el_motivo_de_exclusion(self):
        response = self.client.post(
            reverse("importaciones:plantillas_subir"),
            {"archivo": self._archivo_con_columna_ejercicio_mal_escrita()},
        )
        importacion = Importacion.objects.get()
        hoja = importacion.resultado["hojas"][0]
        self.assertEqual(hoja["items"], [])
        self.assertIsNotNone(hoja["motivo_exclusion"])
        self.assertIn("ejercicio", hoja["motivo_exclusion"])

        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertContains(response, "No se pudo importar")

    def test_preview_no_marca_incluir_por_default_para_hoja_excluida(self):
        response = self.client.post(
            reverse("importaciones:plantillas_subir"),
            {"archivo": self._archivo_con_columna_ejercicio_mal_escrita()},
        )
        importacion = Importacion.objects.get()
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        hoja_formset = response.context["hoja_formset"]
        self.assertFalse(hoja_formset.forms[0].initial["incluir"])

    def test_confirmar_sin_marcar_incluir_no_crea_plantilla_vacia(self):
        response = self.client.post(
            reverse("importaciones:plantillas_subir"),
            {"archivo": self._archivo_con_columna_ejercicio_mal_escrita()},
        )
        importacion = Importacion.objects.get()
        datos = {
            "form-TOTAL_FORMS": "1", "form-INITIAL_FORMS": "1",
            "form-0-nombre_hoja": "Hombres",
            # "incluir" deliberadamente ausente -- checkbox sin marcar, ya
            # que el default ahora es no incluir una hoja sin ejercicios.
            "form-0-objetivo": "Hipertrofia", "form-0-nivel": "principiante",
            "ejercicios-TOTAL_FORMS": "0", "ejercicios-INITIAL_FORMS": "0",
        }
        response = self.client.post(
            reverse("importaciones:plantillas_preview", args=[importacion.pk]), datos,
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(RutinaPlantilla.objects.count(), 0)

    def test_confirmar_incluir_forzado_en_hoja_vacia_falla_en_el_service(self):
        # Defensa en profundidad: aunque alguien arme un POST a mano con
        # `incluir=True` para una hoja de 0 items, el service la rechaza.
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio,
            archivo=self._archivo_con_columna_ejercicio_mal_escrita(),
            usuario=self.staff,
        )
        decisiones = {
            "hojas": [{"nombre_hoja": "Hombres", "incluir": True, "objetivo": "Hipertrofia", "nivel": "principiante"}],
            "ejercicios": {},
        }
        with self.assertRaises(ImportacionInvalida):
            confirmar_importacion_plantillas(
                importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )
        self.assertEqual(RutinaPlantilla.objects.count(), 0)


class AdvertenciasColumnasLlegaAlStaffTests(TestCase):
    """Fix post-review, hallazgo 3: `detectar_columnas` calcula advertencias
    de columnas duplicadas pero antes se descartaban antes de llegar al
    staff (`services.py` hardcodeaba `"advertencias_columnas": []`)."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)
        self.client.login(username="staff", password="clave12345")

    def test_plantillas_columna_ejercicio_duplicada_se_muestra_en_el_preview(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        # Dos columnas "Ejercicio" -- ambas matchean el mismo alias.
        ws.append(["Dia", "Ejercicio", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Press de banca", "otra cosa", 4, "8-12"])

        response = self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        self.assertTrue(importacion.resultado["advertencias_columnas"])

        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertContains(response, "Advertencia")
        self.assertContains(response, "ejercicio")

    def test_biblioteca_columna_nombre_duplicada_se_muestra_en_el_preview(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Dos columnas "Nombre" -- ambas matchean el alias de "nombre".
        ws.append(["Nombre", "Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "otra cosa", "Pecho"])

        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        self.assertTrue(importacion.resultado["advertencias_columnas"])

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "Advertencia")
        self.assertContains(response, "nombre")


class BibliotecaDedupeDentroDelArchivoTests(TestCase):
    """Fix post-review, hallazgo 5: si el MISMO archivo trae dos filas que
    normalizan al mismo nombre (p. ej. "Press de banca" y "PRESS DE
    BANCA"), antes se creaban dos `Ejercicio` -- `Ejercicio` no tiene
    `unique_together`, así que no había ningún otro chequeo que lo evitara."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)
        self.client.login(username="staff", password="clave12345")

    def _archivo_con_duplicado(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Grupo Muscular"])
        ws.append(["Press de banca", "Pecho"])
        ws.append(["PRESS DE BANCA", "Pecho"])  # mismo nombre normalizado
        return _archivo_xlsx(wb)

    def test_previsualizar_deja_solo_la_primera_aparicion_y_lista_la_otra_como_invalida(self):
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo_con_duplicado(), usuario=self.staff,
        )
        self.assertEqual(len(importacion.resultado["items"]), 1)
        self.assertEqual(importacion.resultado["items"][0]["nombre_original"], "Press de banca")
        self.assertEqual(len(importacion.resultado["filas_invalidas"]), 1)
        self.assertIn(
            "duplicado", importacion.resultado["filas_invalidas"][0]["motivo"].lower(),
        )

    def test_confirmar_crea_un_solo_ejercicio_para_el_duplicado(self):
        response = self.client.post(
            reverse("importaciones:biblioteca_subir"), {"archivo": self._archivo_con_duplicado()},
        )
        importacion = Importacion.objects.get()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertContains(response, "duplicado")

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": "{}"},
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Ejercicio.objects.filter(nombre__iexact="press de banca").count(), 1
        )


class EjercicioResolucionMuestraContextoTests(TestCase):
    """Fix post-review, hallazgo 6: la sección "Ejercicios a resolver" del
    preview de plantillas mostraba el nombre NORMALIZADO (lowercase) en vez
    del original, y renderizaba el form completo -- incluyendo
    `ejercicio_existente_id` como un <input type="number"> crudo, sin
    ninguna explicación de qué es esa PK. El spec pide nombre original,
    candidato sugerido y score para los matches ambiguos."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.pecho = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Pecho"
        )
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.staff = User.objects.create_user(username="staff", password="clave12345")
        Perfil.objects.create(usuario=self.staff, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF)
        self.ejercicio_existente = Ejercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Sentadilla",
        )
        self.client.login(username="staff", password="clave12345")

    def test_preview_muestra_nombre_original_candidato_y_score_para_match_ambiguo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hombres"
        ws.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, "Sentadila", 3, "10"])  # typo -> match ambiguo con "Sentadilla"

        self.client.post(
            reverse("importaciones:plantillas_subir"), {"archivo": _archivo_xlsx(wb)},
        )
        importacion = Importacion.objects.get()
        entrada = importacion.resultado["ejercicios_distintos"]["sentadila"]
        self.assertEqual(entrada["tipo"], "ambiguo")

        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        self.assertContains(response, "Sentadila")  # nombre ORIGINAL, no "sentadila"
        self.assertContains(response, entrada["candidato_nombre"])
        self.assertContains(response, str(entrada["score"]))
        # El pk crudo de `ejercicio_existente_id` ya no puede quedar
        # expuesto como un <input type="number"> editable sin etiqueta.
        self.assertNotContains(response, 'type="number"')
        # Una zona de drop por categoría ACTIVA del gimnasio. Antes era un
        # 8 fijo (el tamaño del `TextChoices` global); ahora se cuenta contra
        # el catálogo real, que es distinto en cada gimnasio.
        esperadas = CategoriaEjercicio.objects.for_gimnasio(
            self.gimnasio
        ).filter(activo=True).count()
        self.assertContains(
            response, 'class="rutina-drop-zona"', count=esperadas
        )


class DeteccionTolerantePorContenidoTests(SimpleTestCase):
    """La detección era de coincidencia EXACTA contra la lista de alias, así
    que un encabezado razonable como "NOMBRE DEL EJERCICIO" no matcheaba y el
    archivo entero se descartaba sin explicación."""

    def test_detecta_alias_contenido_en_el_encabezado(self):
        campos, _ = detectar_columnas(
            ["NOMBRE DEL EJERCICIO", "LINK DE YOUTUBE", "CATEGORÍA (grupo)"],
            ALIAS_BIBLIOTECA,
        )
        self.assertEqual(
            campos, {"nombre": 0, "url_video": 1, "grupo_muscular": 2}
        )

    def test_la_coincidencia_exacta_gana_sobre_la_parcial(self):
        """Con 'Ejercicio' (exacto para `nombre`) y 'Nombre del ejercicio'
        (parcial), gana la exacta y la otra no se roba la columna."""
        campos, _ = detectar_columnas(
            ["Nombre del ejercicio", "Ejercicio"], ALIAS_BIBLIOTECA
        )
        self.assertEqual(campos["nombre"], 1)

    def test_una_columna_no_se_asigna_a_dos_campos(self):
        """'Grupo muscular del ejercicio' contiene tanto 'grupo muscular'
        como 'ejercicio'; no puede terminar siendo también la de nombre."""
        campos, _ = detectar_columnas(
            ["Nombre", "Grupo muscular del ejercicio"], ALIAS_BIBLIOTECA
        )
        self.assertEqual(campos["nombre"], 0)
        self.assertEqual(campos["grupo_muscular"], 1)

    def test_gana_el_alias_mas_largo(self):
        campos, _ = detectar_columnas(
            ["Nombre", "Grupo muscular principal"], ALIAS_BIBLIOTECA
        )
        self.assertEqual(campos["grupo_muscular"], 1)

    def test_no_inventa_columnas_que_no_estan(self):
        campos, _ = detectar_columnas(["Nombre", "Observaciones"], ALIAS_BIBLIOTECA)
        self.assertNotIn("grupo_muscular", campos)
        self.assertNotIn("url_video", campos)


class BibliotecaConTituloArribaTests(SimpleTestCase):
    """El caso que ANTES era un error explicado y ahora se importa bien.

    Reproduce la importación #7 del primer cliente pago: el mismo archivo que
    la #8 pero con una fila de título arriba de la tabla. Hasta el 2026-08-31
    `leer_hoja_biblioteca` leía rígido la fila 1, no encontraba la columna
    "nombre" y levantaba `ColumnaRequeridaFaltante` con un mensaje que le
    pedía al staff borrar la fila de arriba. Con `buscar_fila_encabezado` la
    app encuentra la tabla sola.

    NO revertir estos tests a esperar la excepción: el mensaje lindo existía
    porque la app no sabía buscar la tabla, no porque el archivo estuviera
    mal. La propiedad "la app te dice qué leyó" se conserva y se sigue
    fijando en `BibliotecaSinColumnaNombreTests`, para el archivo donde
    realmente no hay tabla en ninguna parte.
    """

    def _hoja(self, filas):
        wb = openpyxl.Workbook()
        ws = wb.active
        for fila in filas:
            ws.append(fila)
        return ws

    def test_un_titulo_arriba_de_la_tabla_ya_no_rompe_la_importacion(self):
        hoja = self._hoja([
            ["Biblioteca de ejercicios 2026", None, None],
            ["NOMBRE", "LINK", "CATEGORÍA"],
            ["Sentadilla", "https://y.com/1", "RODILLA"],
        ])

        items, invalidas, _ = leer_hoja_biblioteca(hoja)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["nombre_original"], "Sentadilla")
        self.assertEqual(items[0]["grupo_muscular_original"], "RODILLA")
        self.assertEqual(invalidas, [])

    def test_la_fila_reportada_es_la_real_de_excel(self):
        """El número que ve el staff tiene que ser el de Excel, no uno
        relativo al encabezado: con el título arriba, la primera fila de
        datos es la 3."""
        hoja = self._hoja([
            ["Biblioteca de ejercicios 2026", None, None],
            ["NOMBRE", "LINK", "CATEGORÍA"],
            ["Sentadilla", "https://y.com/1", "RODILLA"],
        ])

        items, _, _ = leer_hoja_biblioteca(hoja)

        self.assertEqual(items[0]["fila_excel"], 3)

    def test_tambien_con_filas_en_blanco_entre_el_titulo_y_la_tabla(self):
        hoja = self._hoja([
            ["Plan 2026"],
            [],
            [],
            ["NOMBRE", "CATEGORÍA"],
            ["Sentadilla", "RODILLA"],
        ])

        items, _, _ = leer_hoja_biblioteca(hoja)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["fila_excel"], 5)

    def test_el_archivo_bien_armado_sigue_funcionando(self):
        hoja = self._hoja([
            ["NOMBRE", "LINK", "CATEGORÍA"],
            ["Sentadilla", "https://y.com/1", "RODILLA"],
        ])

        items, invalidas, _ = leer_hoja_biblioteca(hoja)

        self.assertEqual(len(items), 1)
        self.assertEqual(invalidas, [])


class BibliotecaSinColumnaNombreTests(SimpleTestCase):
    """Sin columna de nombre en NINGUNA de las filas que se miran, no hay
    nada que importar: sigue siendo un error, no un archivo de cero filas.

    Esta clase cubría antes el caso "título arriba de la tabla", que desde el
    2026-08-31 se importa bien (ver `BibliotecaConTituloArribaTests`). Lo que
    se conserva -- y es lo que de verdad importa -- es que el error le diga al
    staff QUÉ leyó la app, para que no tenga que adivinar.
    """

    def _hoja(self, filas):
        wb = openpyxl.Workbook()
        ws = wb.active
        for fila in filas:
            ws.append(fila)
        return ws

    def test_falta_la_columna_nombre_levanta_error(self):
        hoja = self._hoja([
            ["Resumen mensual", None, None],
            ["Total", "Suma", "Observación"],
            [10, 20, "ok"],
        ])

        with self.assertRaises(ColumnaRequeridaFaltante) as ctx:
            leer_hoja_biblioteca(hoja)

        self.assertEqual(ctx.exception.campo, "nombre")

    def test_el_error_lista_los_encabezados_que_si_encontro(self):
        """Sin esto el staff no tiene forma de saber qué leyó la app: es la
        diferencia entre corregir el archivo y adivinar.

        Se ecoa la fila que MÁS se parece a un encabezado (acá la 2, que al
        menos tiene CATEGORÍA), no la 1: mostrarle el título decorativo del
        archivo no le dice nada.
        """
        hoja = self._hoja([
            ["Resumen mensual", "col b", None],
            ["Total", "CATEGORÍA", "Observación"],
            [10, "RODILLA", "ok"],
        ])

        with self.assertRaises(ColumnaRequeridaFaltante) as ctx:
            leer_hoja_biblioteca(hoja)

        self.assertIn("Total", ctx.exception.encabezados)
        self.assertIn("CATEGORÍA", ctx.exception.encabezados)

    def test_el_error_dice_de_que_fila_salieron_esos_encabezados(self):
        """Decir "la primera fila" dejó de ser cierto cuando la búsqueda pasó
        a mirar las primeras 15."""
        hoja = self._hoja([
            ["Resumen mensual", None, None],
            ["Total", "CATEGORÍA", "Observación"],
            [10, "RODILLA", "ok"],
        ])

        with self.assertRaises(ColumnaRequeridaFaltante) as ctx:
            leer_hoja_biblioteca(hoja)

        self.assertEqual(ctx.exception.fila, 2)

    def test_si_no_hay_nada_parecido_ecoa_la_primera_fila(self):
        """El comportamiento de antes de la búsqueda multi-fila, conservado
        como piso: siempre se muestra algo."""
        hoja = self._hoja([
            ["Resumen mensual", "col b"],
            ["Total", "Suma"],
        ])

        with self.assertRaises(ColumnaRequeridaFaltante) as ctx:
            leer_hoja_biblioteca(hoja)

        self.assertEqual(ctx.exception.fila, 1)
        self.assertIn("Resumen mensual", ctx.exception.encabezados)


class PreviewBibliotecaSinColumnaNombreTests(TestCase):
    """El error de columna faltante tiene que llegar al staff como mensaje,
    no como un preview vacío ni como un 500."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="A", slug="a")
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def _archivo_sin_tabla(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Resumen mensual", None, None])
        ws.append(["Total", "CATEGORÍA", "Observación"])
        ws.append([10, "RODILLA", "ok"])
        return _archivo_xlsx(wb)

    def _archivo_con_titulo_arriba(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Biblioteca de ejercicios 2026", None, None])
        ws.append(["NOMBRE", "LINK", "CATEGORÍA"])
        ws.append(["Sentadilla", "https://y.com/1", "RODILLA"])
        return _archivo_xlsx(wb)

    def test_no_crea_una_importacion_vacia(self):
        with self.assertRaises(ImportacionInvalida):
            previsualizar_importacion_biblioteca(
                gimnasio=self.gimnasio,
                archivo=self._archivo_sin_tabla(),
                usuario=self.usuario,
            )

        self.assertEqual(Importacion.objects.count(), 0)

    def test_el_mensaje_dice_que_columna_falta_y_que_leyo(self):
        with self.assertRaises(ImportacionInvalida) as ctx:
            previsualizar_importacion_biblioteca(
                gimnasio=self.gimnasio,
                archivo=self._archivo_sin_tabla(),
                usuario=self.usuario,
            )

        mensaje = str(ctx.exception)
        self.assertIn("nombre", mensaje)
        self.assertIn("Total", mensaje)
        self.assertIn("15 primeras filas", mensaje)

    def test_el_archivo_con_titulo_arriba_ahora_se_importa(self):
        """Antes del 2026-08-31 este archivo daba `ImportacionInvalida`."""
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio,
            archivo=self._archivo_con_titulo_arriba(),
            usuario=self.usuario,
        )

        self.assertEqual(len(importacion.resultado["items"]), 1)


class ResolverCategoriasTests(SimpleTestCase):
    """Dedupe difuso de nombres de categoría (pedido del dueño: "identificar
    qué palabras, por más que estén mal escritas, quieren decir lo mismo,
    para no crear muchas categorías cuando en realidad son unas pocas").

    El umbral (85) se eligió midiendo `fuzz.ratio` sobre las 12 categorías
    reales del primer cliente más las 8 sembradas por default: el par de
    categorías DISTINTAS más parecido puntúa 61.5 ('Hombros'/'Brazos') y el
    typo que MENOS puntúa entre los que sí deben fusionarse da 88.9
    ('MOVILIDAD'/'MOBILIDAD'). 85 cae en ese hueco de 27 puntos. Los tests de
    abajo fijan los dos bordes.
    """

    CLIENTE = [
        "CORE", "EMPUJE", "ACCESORIOS", "TRACCIÓN", "RODILLA", "CADERA",
        "INTERMITENTE", "DEPORTIVOS", "MUSCLE UP", "MOVILIDAD",
        "SKILLS ANILLAS", "HANDSTAND",
    ]

    def test_catalogo_vacio_crea_una_categoria_por_nombre_distinto(self):
        resueltas = resolver_categorias(self.CLIENTE, {})

        nuevas = {r.nombre for r in resueltas.values() if r.tipo == "nueva"}
        self.assertEqual(len(nuevas), 12)
        self.assertEqual(nuevas, set(self.CLIENTE))

    def test_no_fusiona_categorias_realmente_distintas(self):
        """Regresión del borde de abajo: 'Hombros' y 'Brazos' puntúan 61.5.
        Si alguien baja el umbral, este test lo frena."""
        resueltas = resolver_categorias(["Hombros", "Brazos"], {})

        self.assertNotEqual(
            resueltas["Hombros"].nombre, resueltas["Brazos"].nombre
        )

    def test_fusiona_un_typo_con_la_forma_ya_vista(self):
        resueltas = resolver_categorias(["TRACCIÓN", "TRACION"], {})

        self.assertEqual(
            resueltas["TRACION"].nombre, resueltas["TRACCIÓN"].nombre
        )

    def test_fusiona_el_typo_de_menor_puntaje_del_borde(self):
        """'MOVILIDAD'/'MOBILIDAD' = 88.9, el más flojo de los que deben
        fusionarse. Si alguien sube el umbral, este test lo frena."""
        resueltas = resolver_categorias(["MOVILIDAD", "MOBILIDAD"], {})

        self.assertEqual(
            resueltas["MOBILIDAD"].nombre, resueltas["MOVILIDAD"].nombre
        )

    def test_gana_la_primera_forma_vista_como_nombre_canonico(self):
        resueltas = resolver_categorias(["DEPORTIVOS", "DEPORTIVO"], {})

        self.assertEqual(resueltas["DEPORTIVO"].nombre, "DEPORTIVOS")

    def test_reusa_una_categoria_existente_por_nombre_normalizado(self):
        indice = {"core": (7, "Core")}

        resueltas = resolver_categorias(["CORE"], indice)

        self.assertEqual(resueltas["CORE"].tipo, "existente")
        self.assertEqual(resueltas["CORE"].categoria_id, 7)

    def test_reusa_una_categoria_existente_por_similitud(self):
        """El caso concreto del cliente: su 'CORE' se fusiona con la 'Core'
        que la app siembra por default, en vez de duplicarla."""
        indice = {"core": (7, "Core")}

        resueltas = resolver_categorias(["Coree"], indice)

        self.assertEqual(resueltas["Coree"].tipo, "existente")
        self.assertEqual(resueltas["Coree"].categoria_id, 7)

    def test_el_catalogo_existente_gana_sobre_crear_una_nueva(self):
        indice = {"empuje": (3, "EMPUJE")}

        resueltas = resolver_categorias(["EMPUJES", "EMPUJE"], indice)

        self.assertEqual(resueltas["EMPUJES"].tipo, "existente")
        self.assertEqual(resueltas["EMPUJE"].tipo, "existente")

    def test_una_existente_trae_su_nombre_real_no_el_del_archivo(self):
        """Sin esto el preview muestra el texto crudo del Excel y una fusión
        difusa se lee igual que un match exacto: el staff no tiene cómo
        detectar que "Coree" terminó en "Core" antes de confirmar."""
        indice = {"core": (7, "Core")}

        resueltas = resolver_categorias(["Coree"], indice)

        self.assertEqual(resueltas["Coree"].nombre, "Core")

    def test_ignora_textos_vacios(self):
        resueltas = resolver_categorias(["", None, "   "], {})

        self.assertEqual(resueltas, {})

    def test_las_doce_del_cliente_contra_las_ocho_sembradas(self):
        """La prueba de fuego: el archivo real contra un gimnasio recién
        creado. Solo CORE debe fusionarse con la sembrada; las otras 11 se
        crean, sin que ninguna se coma a otra."""
        indice = {
            "pecho": (1, "Pecho"), "espalda": (2, "Espalda"),
            "piernas": (3, "Piernas"), "hombros": (4, "Hombros"),
            "brazos": (5, "Brazos"), "core": (6, "Core"),
            "cardio": (7, "Cardio"), "cuerpo completo": (8, "Cuerpo completo"),
        }

        resueltas = resolver_categorias(self.CLIENTE, indice)

        existentes = [r for r in resueltas.values() if r.tipo == "existente"]
        nuevas = {r.nombre for r in resueltas.values() if r.tipo == "nueva"}
        self.assertEqual(len(existentes), 1)
        self.assertEqual(existentes[0].categoria_id, 6)
        self.assertEqual(len(nuevas), 11)


class DescartarImportacionesViejasTests(TestCase):
    """`0002_descartar_importaciones_con_formato_viejo`: una importación de
    biblioteca a medio revisar, guardada con el `resultado` de antes de las
    categorías, tiraba 500 al abrir su preview."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def _migrar(self):
        import importlib

        from django.apps import apps

        modulo = importlib.import_module(
            "importaciones.migrations.0002_descartar_importaciones_con_formato_viejo"
        )
        modulo.descartar_en_revision(apps, None)

    def _importacion(self, tipo, estado):
        return Importacion.objects.create(
            gimnasio=self.gimnasio,
            tipo=tipo,
            estado=estado,
            archivo="importaciones/x.xlsx",
            resultado={"items": [], "filas_invalidas": []},
            creado_por=self.usuario,
        )

    def test_descarta_las_de_biblioteca_en_revision(self):
        importacion = self._importacion(
            Importacion.Tipo.BIBLIOTECA, Importacion.Estado.EN_REVISION
        )

        self._migrar()

        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.DESCARTADA)

    def test_no_toca_las_ya_confirmadas(self):
        importacion = self._importacion(
            Importacion.Tipo.BIBLIOTECA, Importacion.Estado.CONFIRMADA
        )

        self._migrar()

        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.CONFIRMADA)

    def test_no_toca_las_de_plantillas(self):
        """El `resultado` de plantillas no cambió de forma."""
        importacion = self._importacion(
            Importacion.Tipo.PLANTILLAS, Importacion.Estado.EN_REVISION
        )

        self._migrar()

        importacion.refresh_from_db()
        self.assertEqual(importacion.estado, Importacion.Estado.EN_REVISION)

    def test_el_preview_de_una_descartada_da_404_no_500(self):
        importacion = self._importacion(
            Importacion.Tipo.BIBLIOTECA, Importacion.Estado.EN_REVISION
        )
        self._migrar()
        self.client.login(username="staff", password="clave-123456")

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

        self.assertEqual(response.status_code, 404)


class BibliotecaPreviewDragYLoteTests(TestCase):
    """El drag-and-drop existía solo en el preview de PLANTILLAS (commit
    5789220). El de biblioteca —el que usa el flujo de "Importar
    ejercicios"— seguía con desplegables sueltos."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )
        self.empuje = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="EMPUJE"
        )
        self.traccion = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="TRACCIÓN"
        )
        self.client.login(username="staff", password="clave-123456")

    def _preview_con_pendientes(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Video"])  # sin columna de categoría a propósito
        ws.append(["Dominadas", "https://y.com/1"])
        ws.append(["Fondos", "https://y.com/2"])
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        return importacion, self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

    def test_hay_una_zona_de_drop_por_categoria_activa(self):
        CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="INACTIVA", activo=False
        )
        _, response = self._preview_con_pendientes()

        self.assertContains(response, 'class="rutina-drop-zona"', count=2)

    def test_cada_pendiente_tiene_su_chip_arrastrable(self):
        _, response = self._preview_con_pendientes()

        self.assertContains(response, 'class="rutina-chip"', count=2)
        self.assertContains(response, 'draggable="true"', count=2)

    def test_la_capa_de_arrastre_esta_oculta_para_lectores_de_pantalla(self):
        """El arrastre nativo no tiene ARIA ni anda en touch: el desplegable
        es el camino accesible y el chip es decoración."""
        _, response = self._preview_con_pendientes()

        self.assertContains(response, 'aria-hidden="true"')

    def test_el_desplegable_sigue_siendo_el_control_autoritativo(self):
        """Si el JS no corre, la pantalla tiene que seguir funcionando."""
        _, response = self._preview_con_pendientes()

        self.assertContains(response, 'class="js-categoria"', count=2)

    def test_hay_controles_de_asignacion_en_lote(self):
        _, response = self._preview_con_pendientes()

        self.assertContains(response, 'id="lote-aplicar"')
        self.assertContains(response, 'id="lote-todos"')
        self.assertContains(response, 'class="js-lote"', count=2)

    def test_las_zonas_llevan_el_id_de_la_categoria_no_su_nombre(self):
        _, response = self._preview_con_pendientes()

        self.assertContains(response, f'data-categoria="{self.empuje.pk}"')

    def test_no_ofrece_categorias_de_otro_gimnasio(self):
        otro = Gimnasio.objects.create(nombre="Otro", slug="otro")
        CategoriaEjercicio.objects.create(gimnasio=otro, nombre="AJENA")

        _, response = self._preview_con_pendientes()

        self.assertNotContains(response, "AJENA")

    def test_muestra_lo_que_decia_el_archivo_para_los_no_resueltos(self):
        """Contexto para decidir: si el Excel decía algo y no se pudo
        resolver, el staff tiene que poder verlo sin abrir el Excel."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Categoría"])
        ws.append(["Dominadas", "ZZZ"])
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        # "ZZZ" no matchea ninguna existente, así que se crea: no queda
        # pendiente. Este test fija que el importador NO deja pendientes
        # cuando el archivo trae categoría.
        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )
        self.assertNotContains(response, 'class="js-lote"')


class SinCategoriaExplicitaTests(TestCase):
    """Elegir "Sin categoría" es una decisión explícita del staff, no un
    default silencioso. Sin esta opción no había forma de confirmar un
    ejercicio cuya categoría todavía no existe (las que va a crear la propia
    importación), ni de importar a un gimnasio con el catálogo vacío."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )
        self.client.login(username="staff", password="clave-123456")

    def _importacion_con_pendiente(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre"])  # sin columna de categoría
        ws.append(["Dominadas"])
        return previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def test_sin_categoria_confirma_y_crea_el_ejercicio(self):
        importacion = self._importacion_con_pendiente()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({"dominadas": {"sin_categoria": True}})},
        )

        self.assertEqual(response.status_code, 302)
        ejercicio = Ejercicio.objects.get(nombre="Dominadas")
        self.assertIsNone(ejercicio.categoria)

    def test_no_elegir_nada_sigue_bloqueando_la_confirmacion(self):
        """La opción explícita no debe convertirse en un default: dejar el
        desplegable sin tocar tiene que seguir frenando el confirm."""
        importacion = self._importacion_con_pendiente()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({})},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Falta resolver")
        self.assertEqual(Ejercicio.objects.count(), 0)

    def test_un_gimnasio_con_el_catalogo_vacio_puede_importar(self):
        """La migración no siembra categorías a un gimnasio sin ejercicios:
        antes de la opción explícita, su desplegable quedaba vacío y la
        importación no se podía confirmar nunca."""
        self.assertEqual(
            CategoriaEjercicio.objects.for_gimnasio(self.gimnasio).count(), 0
        )
        importacion = self._importacion_con_pendiente()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({"dominadas": {"sin_categoria": True}})},
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Ejercicio.objects.count(), 1)

    def test_categoria_id_booleano_no_pasa_como_pk(self):
        """`isinstance(True, int)` es True en Python: sin el guard de bool,
        `categoria_id: true` se colaba como si fuera el pk 1."""
        importacion = self._importacion_con_pendiente()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({"dominadas": {"categoria_id": True}})},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Categoría inválida")


class CategoriasACrearCuentaSoloLasRealesTests(TestCase):
    """El resumen del preview contaba TODAS las filas parseadas, incluidas
    las descartadas por duplicadas y las de ejercicios que ya existen --
    ninguna de las dos llega a crear una categoría."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def _archivo(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Categoría"])
        ws.append(["Dominadas", "TRACCIÓN"])
        return _archivo_xlsx(wb)

    def test_subir_el_mismo_archivo_dos_veces_no_anuncia_categorias_de_mas(self):
        primera = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )
        self.assertEqual(primera.resultado["categorias_a_crear"], ["TRACCIÓN"])
        confirmar_importacion_biblioteca(
            importacion=primera,
            gimnasio=self.gimnasio,
            decisiones={"items": {"dominadas": {"incluir": True}}},
        )

        segunda = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=self._archivo(), usuario=self.usuario,
        )

        # El ejercicio ya existe: no se recrea, y su categoría tampoco.
        self.assertEqual(segunda.resultado["categorias_a_crear"], [])


class ImportacionBibliotecaLargosTests(TestCase):
    """Una celda demasiado larga no puede voltear la importación entera.

    En Postgres un `varchar` desbordado es un `DataError` que aborta la
    transacción: el Excel real de un cliente traía dos links de 306
    caracteres y eso habría dejado sin importar los otros 746 ejercicios,
    con un 500 sin explicación. SQLite no valida largos, así que estos tests
    chequean el guard directamente, no el error de la base."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user(username="staff", password="clave12345")

    def _preview(self, filas):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Link"])
        for fila in filas:
            ws.append(fila)
        return previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def test_un_link_largo_pero_valido_entra(self):
        """306 caracteres es lo que traía el archivo real: tiene que entrar,
        no descartarse. Por eso `url_video` es `max_length=500`."""
        link = "https://www.google.com/search?q=" + "a" * 274
        self.assertEqual(len(link), 306)
        importacion = self._preview([["Puente supino", link]])
        self.assertEqual(len(importacion.resultado["items"]), 1)
        self.assertEqual(importacion.resultado["items"][0]["url_video"], link)

    def test_un_link_imposible_descarta_solo_esa_fila(self):
        importacion = self._preview([
            ["Puente supino", "https://x.com/" + "a" * 600],
            ["Sentadilla", "https://youtu.be/ok"],
        ])
        self.assertEqual(len(importacion.resultado["items"]), 1)
        self.assertEqual(importacion.resultado["items"][0]["nombre_original"], "Sentadilla")
        self.assertEqual(len(importacion.resultado["filas_invalidas"]), 1)
        self.assertIn("link del video", importacion.resultado["filas_invalidas"][0]["motivo"])

    def test_un_nombre_imposible_descarta_solo_esa_fila(self):
        importacion = self._preview([
            ["N" * 200, ""],
            ["Sentadilla", ""],
        ])
        self.assertEqual(len(importacion.resultado["items"]), 1)
        self.assertIn("nombre", importacion.resultado["filas_invalidas"][0]["motivo"])


class ImportacionBibliotecaEscalaTests(TestCase):
    """El confirm de biblioteca tiene que costar lo mismo con 20 filas que
    con 200: es el flujo que un gimnasio real usa una sola vez, con toda su
    biblioteca (748 ejercicios en el caso que lo rompió)."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.categoria = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Empuje"
        )
        self.usuario = User.objects.create_user(username="staff", password="clave12345")

    def _importacion(self, cantidad):
        items = [
            {
                "fila_excel": i + 2,
                "nombre_original": f"Ejercicio {i}",
                "nombre_normalizado": f"ejercicio {i}",
                "grupo_muscular_original": "Empuje",
                "url_video": "",
                "categoria_resuelta": {
                    "tipo": "existente",
                    "categoria_id": self.categoria.pk,
                    "nombre": self.categoria.nombre,
                },
                "match": {"tipo": "nuevo"},
            }
            for i in range(cantidad)
        ]
        return Importacion.objects.create(
            gimnasio=self.gimnasio,
            tipo=Importacion.Tipo.BIBLIOTECA,
            resultado={"items": items, "filas_invalidas": [], "advertencias_columnas": []},
            creado_por=self.usuario,
        )

    def _queries_para(self, cantidad):
        importacion = self._importacion(cantidad)
        decisiones = {"items": {
            f"ejercicio {i}": {
                "incluir": True, "categoria_id": None, "sin_categoria": False,
            }
            for i in range(cantidad)
        }}
        with CaptureQueriesContext(connection) as consultas:
            creados = confirmar_importacion_biblioteca(
                importacion=importacion, gimnasio=self.gimnasio, decisiones=decisiones,
            )
        self.assertEqual(len(creados), cantidad)
        return len(consultas)

    def test_el_costo_en_queries_no_crece_con_la_cantidad_de_filas(self):
        """Regresión del 502 del 2026-08-27. El confirm hacía DOS queries por
        fila -- el SELECT de la categoría en `_categoria_para` y el INSERT de
        `Ejercicio.objects.create()` -- o sea ~1500 round-trips contra Neon
        para un archivo de 748 ejercicios: más de los 30 s de timeout de
        gunicorn, worker muerto y 502 en la cara del staff.

        Se comparan dos tamaños en vez de fijar un número exacto (mismo
        criterio que el test de `select_related` de `alumnos:accesos`): un
        `assertNumQueries` literal se rompe con cualquier cambio interno de
        Django sin que haya una regresión real."""
        chico = self._queries_para(20)
        grande = self._queries_para(200)
        self.assertLessEqual(
            grande, chico + 5,
            f"El confirm escala con la cantidad de filas: {chico} queries con "
            f"20 ejercicios, {grande} con 200.",
        )


class CategoriasNuevasSeOfrecenEnElPreviewTests(TestCase):
    """Reporte real del 2026-08-27 (gimnasio "Vida Plena", capturas): el
    Excel traía 12 categorías nuevas y una fila con la celda de categoría
    vacía. Esa fila quedaba pendiente y el desplegable solo ofrecía
    «Sin categoría» -- las 12 categorías del propio archivo no estaban,
    porque todavía no existen en la base. Con el catálogo del gimnasio
    vacío (el caso de un primer import) el staff no tenía NINGUNA categoría
    donde ubicar el ejercicio."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )
        self.client.login(username="staff", password="clave-123456")

    def _importacion(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Categoría"])
        # El pendiente va PRIMERO a propósito: si otro ejercicio crea antes
        # la categoría que el staff elige, `crear_o_reusar` la encuentra ya
        # cacheada y el nombre que mandó el POST nunca llega a escribirse --
        # el orden del archivo decide si el agujero se ejercita o no.
        ws.append(["Press Pallof estático", None])  # la celda vacía del archivo real
        ws.append(["Sentadilla búlgara", "RODILLA"])
        ws.append(["Remo con barra", "TRACCIÓN"])
        ws.append(["Muscle up", "SKILLS ANILLAS"])
        return previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def test_el_desplegable_ofrece_las_categorias_que_crea_esta_importacion(self):
        importacion = self._importacion()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

        self.assertContains(response, 'value="nueva:RODILLA"')
        self.assertContains(response, 'value="nueva:TRACCIÓN"')

    def test_hay_una_zona_de_drop_por_cada_categoria_nueva(self):
        importacion = self._importacion()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

        self.assertContains(response, 'class="rutina-drop-zona"', count=3)

    def test_la_zona_de_drop_marca_las_que_todavia_no_existen(self):
        """Un gimnasio con "CORE" que importa "CORE FUNCIONAL" ve dos zonas
        que se leen igual: sin la marca no hay forma de saber cuál existe."""
        importacion = self._importacion()

        response = self.client.get(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk])
        )

        self.assertContains(response, "RODILLA (nueva)</div>")

    def test_confirmar_con_una_categoria_nueva_la_reusa_en_vez_de_duplicarla(self):
        importacion = self._importacion()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps(
                {"press pallof estatico": {"categoria_nueva": "RODILLA"}}
            )},
        )

        self.assertEqual(response.status_code, 302)
        ejercicio = Ejercicio.objects.get(nombre="Press Pallof estático")
        self.assertEqual(ejercicio.categoria.nombre, "RODILLA")
        self.assertEqual(
            CategoriaEjercicio.objects.for_gimnasio(self.gimnasio)
            .filter(nombre_normalizado="rodilla").count(),
            1,
        )

    def test_una_categoria_que_el_archivo_no_traia_no_se_crea_desde_el_post(self):
        """El nombre viaja como texto en el POST: tiene que validarse contra
        lo que ESTA importación anunció, no crearse a ciegas."""
        importacion = self._importacion()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps(
                {"press pallof estatico": {"categoria_nueva": "INVENTADA"}}
            )},
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(
            CategoriaEjercicio.objects.filter(nombre_normalizado="inventada").exists()
        )
        self.assertEqual(Ejercicio.objects.count(), 0)

    def test_el_nombre_que_se_guarda_es_el_del_archivo_no_el_del_post(self):
        """`normalizar_texto` colapsa los espacios internos pero
        `CategoriaEjercicio.save()` solo hace `.strip()`: un nombre con 100
        espacios en el medio pasaba el guard (normaliza igual) y se escribía
        con 113 caracteres en un `varchar(60)`. En Postgres eso es un
        `DataError` que voltea la transacción entera; SQLite no lo valida, así
        que el test mira el largo y el nombre, no la excepción."""
        importacion = self._importacion()

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps({
                "press pallof estatico": {
                    "categoria_nueva": "SKILLS" + " " * 100 + "ANILLAS"
                }
            })},
        )

        self.assertEqual(response.status_code, 302)
        categoria = Ejercicio.objects.get(nombre="Press Pallof estático").categoria
        self.assertEqual(categoria.nombre, "SKILLS ANILLAS")
        self.assertLessEqual(
            len(categoria.nombre),
            CategoriaEjercicio._meta.get_field("nombre").max_length,
        )

    def test_elegirla_en_minusculas_no_cambia_como_se_muestra(self):
        """Variante menor del mismo agujero: el catálogo creado tiene que
        coincidir con lo que el preview anunció."""
        importacion = self._importacion()

        self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            {"resoluciones": json.dumps(
                {"press pallof estatico": {"categoria_nueva": "rodilla"}}
            )},
        )

        self.assertEqual(
            Ejercicio.objects.get(nombre="Press Pallof estático").categoria.nombre,
            "RODILLA",
        )

    def test_un_post_rechazado_no_borra_lo_que_el_staff_ya_habia_elegido(self):
        """Las decisiones viven en el blob JSON, no en el HTML: sin
        devolverlas, un único pendiente olvidado en un archivo de 748 filas
        hace que el staff tenga que rehacer TODAS las demás."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Categoría"])
        ws.append(["Sentadilla búlgara", "RODILLA"])
        ws.append(["Press Pallof estático", None])
        ws.append(["Plancha lateral", None])
        importacion = previsualizar_importacion_biblioteca(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

        response = self.client.post(
            reverse("importaciones:biblioteca_preview", args=[importacion.pk]),
            # Resuelve uno solo: el otro dispara "Falta resolver".
            {"resoluciones": json.dumps(
                {"press pallof estatico": {"categoria_nueva": "RODILLA"}}
            )},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Falta resolver")
        self.assertContains(response, 'id="resoluciones-previas"')
        self.assertContains(response, "press pallof estatico")
        self.assertContains(response, "categoria_nueva")


class BuscarFilaEncabezadoTests(SimpleTestCase):
    """La fila de títulos no siempre es la 1.

    La planilla real del primer cliente pago la tiene en la fila 12: arriba
    hay logo, objetivo, fechas de inicio/fin y un "Cumplim: 25%". Hasta acá
    los tres lectores hacían `next(ws.iter_rows(min_row=1, max_row=1))` y el
    archivo entero se rechazaba.
    """

    REQUERIDOS = ("ejercicio", "series", "repeticiones")

    def _hoja(self, filas):
        wb = openpyxl.Workbook()
        ws = wb.active
        for fila in filas:
            ws.append(fila)
        return ws

    def test_encabezado_en_la_primera_fila(self):
        ws = self._hoja([
            ["Ejercicio", "Series", "Repeticiones"],
            ["Sentadilla", 4, "10"],
        ])
        encabezado = buscar_fila_encabezado(ws, ALIAS_PLANTILLA, self.REQUERIDOS)
        self.assertEqual(encabezado.fila, 1)
        self.assertEqual(encabezado.campos["ejercicio"], 0)

    def test_encabezado_despues_de_un_titulo_y_filas_en_blanco(self):
        ws = self._hoja([
            ["Plan de entrenamiento 2026"],
            [],
            ["Alumna:", "Eve Colazo"],
            [],
            ["Ejercicio", "Series", "Repeticiones"],
            ["Sentadilla", 4, "10"],
        ])
        encabezado = buscar_fila_encabezado(ws, ALIAS_PLANTILLA, self.REQUERIDOS)
        self.assertEqual(encabezado.fila, 5)
        self.assertEqual(encabezado.campos["series"], 1)

    def test_sin_ninguna_fila_de_titulos_devuelve_none(self):
        ws = self._hoja([
            ["Total", "Suma", "Observación"],
            [10, 20, "ok"],
        ])
        self.assertIsNone(
            buscar_fila_encabezado(ws, ALIAS_PLANTILLA, self.REQUERIDOS)
        )

    def test_una_fila_de_datos_no_se_confunde_con_el_encabezado(self):
        """El guardarraíl de la búsqueda multi-fila: si una fila de datos
        pudiera hacerse pasar por encabezado, el parser leería columnas
        corridas y produciría basura plausible en vez de un error."""
        ws = self._hoja([
            ["Dia", "Series", "Repeticiones"],
            [1, 4, "8-12"],
            [1, 3, "10"],
        ])
        # Falta `ejercicio` en TODA la hoja: ninguna fila califica.
        self.assertIsNone(
            buscar_fila_encabezado(ws, ALIAS_PLANTILLA, self.REQUERIDOS)
        )

    def test_no_mira_mas_alla_de_la_ventana(self):
        filas = [["ruido"]] * 20 + [["Ejercicio", "Series", "Repeticiones"]]
        ws = self._hoja(filas)
        self.assertIsNone(
            buscar_fila_encabezado(ws, ALIAS_PLANTILLA, self.REQUERIDOS, max_filas=15)
        )
        self.assertEqual(
            buscar_fila_encabezado(
                ws, ALIAS_PLANTILLA, self.REQUERIDOS, max_filas=25
            ).fila,
            21,
        )

    def test_mejor_encabezado_parcial_sirve_para_el_mensaje_de_error(self):
        """Cuando no se encuentra la tabla, al staff hay que decirle qué SÍ
        se leyó. Gana la fila con más campos detectados, no la fila 1."""
        ws = self._hoja([
            ["Biblioteca de ejercicios 2026"],
            ["Nombre", "Categoría"],
            ["Sentadilla", "Piernas"],
        ])
        parcial = mejor_encabezado_parcial(ws, ALIAS_BIBLIOTECA)
        self.assertEqual(parcial.fila, 2)
        self.assertIn("Nombre", parcial.valores)

    def test_mejor_encabezado_parcial_cae_en_la_fila_1_si_no_hay_nada(self):
        ws = self._hoja([["Total", "Suma"], [1, 2]])
        parcial = mejor_encabezado_parcial(ws, ALIAS_BIBLIOTECA)
        self.assertEqual(parcial.fila, 1)
        self.assertEqual(parcial.valores, ["Total", "Suma"])


def _hoja_matriz_ancha(*, filas_extra=(), semanas=("SEMANA 1", "SEMANA 2"),
                       fila_inicio=1, subcampos=("Series", "Reps", "Carga", "RPE")):
    """Reproduce la forma de la planilla real: dos filas de encabezado (grupos
    combinados arriba, subcampos abajo), el día en la columna A y el código de
    bloque + el nombre en B y C."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(fila_inicio - 1):
        ws.append(["Powered by Simplify Trainers"])

    grupos = [None, " EJERCICIOS", None, "Videos"]
    subs = [None, None, None, None]
    for semana in semanas:
        grupos.append(semana)
        grupos.extend([None] * (len(subcampos) - 1))
        subs.extend(subcampos)
    ws.append(grupos)
    ws.append(subs)

    # Las tres filas base no son decorativas: `MIN_FILAS_CON_NOMBRE` exige al
    # menos tres ejercicios con nombre real para aceptar el layout, así que un
    # fixture más chico se rechazaría (y con razón: no sería un plan).
    # `filas_extra` se SUMA a estas, no las reemplaza.
    for fila in (
        ["DÍA 1\n• CORE", "A1.", "Plancha", None, 4, "20", None, None, 4, "25", None, None],
        [None, "A2.", "Press Pallof", None, 3, "12", "10KG", None, 3, "15", "10KG", None],
        ["DÍA 2\n• TREN SUPERIOR", "A1.", "Remo en TRX", None, 4, "10", None, None, 4, "12", None, None],
    ):
        ws.append(list(fila))
    for fila in filas_extra:
        ws.append(list(fila))
    return ws


class DeteccionLayoutTests(SimpleTestCase):
    """Cuál de los dos lectores se elige, y por qué el orden importa."""

    def test_una_hoja_ancha_se_detecta_como_ancha(self):
        self.assertIsNotNone(detectar_matriz_ancha(_hoja_matriz_ancha()))

    def test_una_hoja_ancha_con_el_encabezado_abajo_tambien(self):
        enc = detectar_matriz_ancha(_hoja_matriz_ancha(fila_inicio=12))
        self.assertIsNotNone(enc)
        self.assertEqual(enc.fila_grupos, 12)
        self.assertEqual(enc.fila_subcampos, 13)

    def test_una_hoja_larga_no_se_confunde_con_ancha(self):
        """`Semana` a secas no matchea: el regex exige el dígito. Sin eso, el
        layout de siempre caería en el lector nuevo."""
        self.assertIsNone(detectar_matriz_ancha(_hoja_plantilla_basica()))

    def test_una_hoja_ancha_nunca_cae_al_lector_largo(self):
        """El riesgo central del diseño: la fila de subcampos tiene
        Series/Reps/Carga y la de grupos tiene EJERCICIOS, así que el lector
        largo la aceptaría y produciría filas plausibles con las columnas
        corridas. Basura silenciosa es peor que cero items."""
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        self.assertGreater(len(hoja.items), 0)
        self.assertEqual({i.semana for i in hoja.items}, {1, 2})

    def test_una_hoja_auxiliar_no_se_detecta(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Powered by Simplify Trainers"])
        ws.append(["NOMBRE:", "EVE COLAZO"])
        ws.append(["OBJETIVO:", "Fuerza"])
        self.assertIsNone(detectar_matriz_ancha(ws))

    def test_una_sola_semana_no_alcanza(self):
        """Con un único `SEMANA 1` no hay matriz: puede ser el título de una
        tabla larga cualquiera."""
        self.assertIsNone(detectar_matriz_ancha(_hoja_matriz_ancha(semanas=("SEMANA 1",))))

    def test_una_tabla_resumen_por_semanas_no_genera_ejercicios_fantasma(self):
        """Guarda anti-falso-positivo: una hoja de progreso con columnas
        SEMANA 1/SEMANA 2 pasa los primeros pasos, pero no tiene una columna
        de nombres de ejercicio con contenido real."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "SEMANA 1", None, "SEMANA 2", None])
        ws.append([None, "Series", "Reps", "Series", "Reps"])
        ws.append(["Total", 10, "20", 12, "22"])
        self.assertIsNone(detectar_matriz_ancha(ws))

    def test_la_deteccion_no_escanea_la_hoja_entera(self):
        """Una hoja auxiliar de miles de filas tiene que descartarse barato."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(3000):
            ws.append(["ruido", 1, 2, 3])

        filas_leidas = set()
        original = ws.cell

        def espia(row=None, column=None, **kwargs):
            filas_leidas.add(row)
            return original(row=row, column=column, **kwargs)

        ws.cell = espia
        self.assertIsNone(detectar_matriz_ancha(ws))
        self.assertLessEqual(max(filas_leidas), 40)


class LeerHojaAnchaTests(SimpleTestCase):
    """Una fila de Excel produce un item POR SEMANA."""

    def test_emite_un_item_por_ejercicio_y_semana(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        self.assertEqual(len(hoja.items), 6)  # 3 ejercicios x 2 semanas

    def test_el_dia_se_arrastra_hacia_abajo(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        por_nombre = {i.ejercicio_original: i.dia for i in hoja.items}
        self.assertEqual(por_nombre["Plancha"], 1)
        self.assertEqual(por_nombre["Press Pallof"], 1)  # sin marcador propio
        self.assertEqual(por_nombre["Remo en TRX"], 2)

    def test_el_nombre_del_dia_sale_del_mismo_marcador(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        nombres = {i.dia: i.dia_nombre for i in hoja.items}
        self.assertEqual(nombres[1], "CORE")
        self.assertEqual(nombres[2], "TREN SUPERIOR")

    def test_gana_la_celda_de_dia_con_mas_texto(self):
        """En la planilla real el marcador aparece repetido por los merges:
        `DÍA 2` pelado en una columna y `DÍA 2 + descripción` en otra. Quedarse
        con el primero perdería el nombre del día."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, " EJERCICIOS", None, "SEMANA 1", None, None, None, "SEMANA 2", None, None, None])
        ws.append([None, None, None, None, "Series", "Reps", "Carga", "RPE", "Series", "Reps", "Carga", "RPE"])
        # El `DÍA 1` pelado de la columna A viene ANTES que la celda rica de la
        # columna B: si ganara el primero que aparece, se perdería el nombre.
        ws.append(["DÍA 1", "DÍA 1\n• CORE\n• MOVILIDAD", "A1.", "Plancha lateral", 4, "20", None, None, 4, "25", None, None])
        ws.append([None, None, "A2.", "Press Pallof", 4, "20", None, None, 4, "25", None, None])
        ws.append([None, None, "A3.", "Remo invertido", 4, "20", None, None, 4, "25", None, None])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items[0].dia_nombre, "CORE · MOVILIDAD")

    def test_guarda_el_codigo_de_bloque(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        self.assertEqual(hoja.items[0].bloque, "A1")

    def test_el_rpe_del_archivo_se_descarta(self):
        """El RPE de la app lo carga el alumno sobre SU rutina asignada. El
        del Excel es de otra persona y de un ciclo cerrado."""
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            [None, "B1.", "Sentadilla búlgara", None,
             4, "20", None, "🟡 Podría seguir con esta intensidad",
             4, "25", None, "⚫ Debería bajar la intensidad"],
        )))
        for item in hoja.items:
            self.assertNotIn("Podría seguir", item.notas)
            self.assertNotIn("bajar", item.notas)

    def test_una_semana_sin_datos_no_emite_item(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            [None, "B1.", "Solo primera semana", None, 4, "20", None, None, None, None, None, None],
        )))
        solo = [i for i in hoja.items if i.ejercicio_original == "Solo primera semana"]
        self.assertEqual(len(solo), 1)
        self.assertEqual(solo[0].semana, 1)

    def test_una_fila_sin_nombre_ni_datos_se_saltea_en_silencio(self):
        """La planilla real trae slots vacíos: el código de bloque cargado
        (`D3.`) y el resto en blanco. No es un error que valga reportar."""
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            [None, "D3.", None, None, None, None, None, None, None, None, None, None],
        )))
        self.assertEqual(len(hoja.items), 6)  # las 3 filas base x 2 semanas
        self.assertEqual(hoja.filas_invalidas, [])

    def test_una_fila_con_datos_pero_sin_nombre_si_se_reporta(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            [None, "D3.", None, None, 4, "10", None, None, None, None, None, None],
        )))
        self.assertEqual(len(hoja.filas_invalidas), 1)
        self.assertIn("nombre", hoja.filas_invalidas[0].motivo)

    def test_series_no_numerica_solo_invalida_esa_semana(self):
        """La diferencia de cardinalidad con el layout largo: una fila de
        Excel son hasta 4 items, así que un dato malo en la semana 2 no puede
        llevarse puestas las otras tres."""
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            [None, "B1.", "Peso muerto rumano", None, 4, "20", None, None, "cuatro", "25", None, None],
        )))
        muerto = [i for i in hoja.items if i.ejercicio_original == "Peso muerto rumano"]
        self.assertEqual(len(muerto), 1)
        self.assertEqual(muerto[0].semana, 1)
        self.assertEqual(len(hoja.filas_invalidas), 1)
        self.assertIn("Semana 2", hoja.filas_invalidas[0].motivo)

    def test_dias_por_semana_usa_el_dia_mas_alto_igual_que_el_lector_largo(self):
        """No `len(set(...))`: con días 1, 2 y 4 el plan tiene 4 días, no 3.
        Los dos layouts tienen que coincidir en el criterio."""
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        self.assertEqual(hoja.dias_por_semana, 2)

        con_salto = leer_hoja_plantilla(_hoja_matriz_ancha(filas_extra=(
            ["DÍA 4", "A1.", "Sentadilla frontal", None, 4, "10", None, None, 4, "12", None, None],
        )))
        self.assertEqual(con_salto.dias_por_semana, 4)

    def test_el_orden_es_secuencial_dentro_de_cada_dia_y_semana(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        dia1_sem1 = [i for i in hoja.items if i.dia == 1 and i.semana == 1]
        self.assertEqual([i.orden for i in dia1_sem1], [1, 2])

    def test_la_carga_llega_a_kilos(self):
        hoja = leer_hoja_plantilla(_hoja_matriz_ancha())
        press = [i for i in hoja.items if i.ejercicio_original == "Press Pallof"]
        self.assertEqual(press[0].kilos, "10KG")

    def test_ancha_sin_ejercicios_se_excluye_con_motivo(self):
        """Nunca una hoja vacía y muda: si se detectó la matriz pero no salió
        nada, hay que decir por qué."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, None, " EJERCICIOS", "SEMANA 1", None, "SEMANA 2", None])
        ws.append([None, None, None, "Series", "Reps", "Series", "Reps"])
        ws.append(["DÍA 1", "A1.", "Plancha lateral", None, None, None, None])
        ws.append([None, "A2.", "Remo invertido", None, None, None, None])
        ws.append([None, "A3.", "Press militar", None, None, None, None])
        hoja = leer_hoja_plantilla(ws)
        self.assertEqual(hoja.items, [])
        self.assertIsNotNone(hoja.motivo_exclusion)
        self.assertIn("semanas", hoja.motivo_exclusion)


class SinonimosDeTerminologiaTests(SimpleTestCase):
    """Cada entrenador nombra las cosas distinto y el importador tiene que
    adaptarse a él, no al revés.

    "Microciclo" es semana, "sesión" es día, "carga" es kilos. El vocabulario
    va en ES y EN porque las planillas compradas suelen venir mezcladas.
    """

    def _hoja(self, encabezados):
        # semana=2 y dia=3, NO 1: si la columna no se detecta, el parser cae a
        # los defaults (semana 1, día 1) y un test con 1 pasaría sin probar
        # nada.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(encabezados)
        ws.append([2, 3, "Sentadilla", 4, "10", "90s", ""])
        return ws

    def test_microciclo_es_semana(self):
        for palabra in ["Microciclo", "MICRO", "Week", "Wk", "Sem"]:
            with self.subTest(palabra=palabra):
                hoja = leer_hoja_plantilla(self._hoja(
                    [palabra, "Dia", "Ejercicio", "Series", "Repeticiones", "Descanso", "Notas"]
                ))
                self.assertIsNone(hoja.motivo_exclusion)
                self.assertEqual(hoja.items[0].semana, 2)

    def test_sesion_es_dia(self):
        for palabra in ["Sesión", "Sesion", "Session", "Jornada", "Day"]:
            with self.subTest(palabra=palabra):
                hoja = leer_hoja_plantilla(self._hoja(
                    ["Semana", palabra, "Ejercicio", "Series", "Repeticiones", "Descanso", "Notas"]
                ))
                self.assertIsNone(hoja.motivo_exclusion)
                self.assertEqual(hoja.items[0].dia, 3)

    def test_variantes_de_los_campos_numericos(self):
        casos = [
            (["Semana", "Dia", "Movement", "Sets", "Repetitions", "Rest", "Notes"], "en"),
            (["Semana", "Dia", "Movimiento", "Serie", "Repes", "Pausa", "Observaciones"], "es"),
        ]
        for encabezados, idioma in casos:
            with self.subTest(idioma=idioma):
                hoja = leer_hoja_plantilla(self._hoja(encabezados))
                self.assertEqual(len(hoja.items), 1, hoja.motivo_exclusion)
                self.assertEqual(hoja.items[0].semana, 2)
                self.assertEqual(hoja.items[0].dia, 3)
                self.assertEqual(hoja.items[0].series, 4)
                self.assertEqual(hoja.items[0].repeticiones, "10")
                self.assertEqual(hoja.items[0].descanso, "90s")

    def test_variantes_de_carga(self):
        for palabra in ["Carga", "Peso", "Kg", "Kgs", "Load", "Weight"]:
            with self.subTest(palabra=palabra):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Ejercicio", "Series", "Repeticiones", palabra])
                ws.append(["Sentadilla", 4, "10", "20KG"])
                hoja = leer_hoja_plantilla(ws)
                self.assertEqual(hoja.items[0].kilos, "20KG")


class LargosDePlantillaTests(TestCase):
    """Postgres rechaza un varchar largo con un DataError que voltea la
    transacción entera; SQLite no valida nada, así que esto NUNCA lo iba a
    encontrar un test local sin chequearlo a mano (ISSUES.md 2026-08-27, el
    link de 306 caracteres). La fila se descarta en el PREVIEW, con motivo,
    no en el INSERT.
    """

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="A", slug="a")
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def _archivo(self, filas):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Semana", "Dia", "Ejercicio", "Series", "Repeticiones", "Kilos"])
        for fila in filas:
            ws.append(fila)
        return _archivo_xlsx(wb)

    def _hojas(self, filas):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=self._archivo(filas), usuario=self.usuario
        )
        return importacion.resultado["hojas"][0]

    def test_una_fila_que_entra_no_se_toca(self):
        hoja = self._hojas([[1, 1, "Sentadilla", 4, "10", "20KG"]])
        self.assertEqual(len(hoja["items"]), 1)
        self.assertEqual(hoja["filas_invalidas"], [])

    def test_repeticiones_demasiado_largas_descartan_solo_esa_fila(self):
        hoja = self._hojas([
            [1, 1, "Sentadilla", 4, "10", "20KG"],
            [1, 1, "Press", 4, "x" * 40, "20KG"],
        ])
        self.assertEqual(len(hoja["items"]), 1)
        self.assertEqual(len(hoja["filas_invalidas"]), 1)
        self.assertIn("repeticiones", hoja["filas_invalidas"][0]["motivo"].lower())

    def test_un_nombre_de_ejercicio_larguisimo_tambien(self):
        hoja = self._hojas([
            [1, 1, "Sentadilla", 4, "10", "20KG"],
            [1, 1, "N" * 200, 4, "10", "20KG"],
        ])
        self.assertEqual(len(hoja["items"]), 1)
        self.assertEqual(len(hoja["filas_invalidas"]), 1)

    def test_el_ejercicio_descartado_no_queda_pendiente_de_resolucion(self):
        """Si el nombre siguiera en `ejercicios_distintos`, el preview le
        pediría al staff clasificar un ejercicio que no se va a crear."""
        hoja = self._hojas([
            [1, 1, "Sentadilla", 4, "10", "20KG"],
            [1, 1, "N" * 200, 4, "10", "20KG"],
        ])
        importacion = Importacion.objects.get()
        self.assertNotIn(
            normalizar_texto("N" * 200), importacion.resultado["ejercicios_distintos"]
        )

    def test_una_semana_fuera_del_ciclo_se_descarta(self):
        """`bulk_create` no corre validadores, así que sin este chequeo una
        semana 5 entraría a la base saltándose el MaxValueValidator."""
        hoja = self._hojas([
            [1, 1, "Sentadilla", 4, "10", "20KG"],
            [5, 1, "Press", 4, "10", "20KG"],
        ])
        self.assertEqual(len(hoja["items"]), 1)
        self.assertEqual(len(hoja["filas_invalidas"]), 1)
        self.assertIn("semana", hoja["filas_invalidas"][0]["motivo"].lower())


class FilaExcelReportadaTests(TestCase):
    """El número de fila que se le muestra al staff tiene que ser el de Excel.

    Reportar `orden` (la posición dentro del día) mandaba al entrenador a
    buscar una celda que no existe.
    """

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="A", slug="a")
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def test_la_fila_descartada_se_reporta_con_su_numero_de_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Semana", "Dia", "Ejercicio", "Series", "Repeticiones"])
        ws.append([1, 1, "Sentadilla", 4, "10"])       # fila 2
        ws.append([1, 1, "Press", 4, "10"])            # fila 3
        ws.append([5, 1, "Fuera de ciclo", 4, "10"])   # fila 4, se descarta

        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario
        )
        invalidas = importacion.resultado["hojas"][0]["filas_invalidas"]

        self.assertEqual(len(invalidas), 1)
        self.assertEqual(invalidas[0]["fila_excel"], 4)

    def test_tambien_en_la_matriz_ancha(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Titulo del plan"])
        ws.append([None, " EJERCICIOS", None, "SEMANA 1", None, "SEMANA 2", None])
        ws.append([None, None, None, "Series", "Reps", "Series", "Reps"])
        ws.append(["DÍA 1", "A1.", "Plancha lateral", 4, "20", 4, "25"])   # fila 4
        ws.append([None, "A2.", "Press Pallof", 4, "20", 4, "25"])          # fila 5
        ws.append([None, "A3.", "Remo invertido", 4, "20", 4, "x" * 40])    # fila 6

        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario
        )
        invalidas = importacion.resultado["hojas"][0]["filas_invalidas"]

        self.assertEqual(len(invalidas), 1)
        self.assertEqual(invalidas[0]["fila_excel"], 6)


class SeleccionDeHojasTests(TestCase):
    """El archivo real del primer cliente pago trae 7 hojas y 6 son
    auxiliares (`AUX` con 3206 filas, `Movilidad Articular` con 1020,
    `Avatar`, `Logros`, `Carga de Datos`, `Plantilla - aux`).

    Sin este paso el preview mostraba las 7 y el staff tenía que destildar 6
    para llegar a la única que le importaba.
    """

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.piernas = CategoriaEjercicio.objects.create(
            gimnasio=self.gimnasio, nombre="Piernas"
        )
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )
        self.client.login(username="staff-a", password="clave-123456")

        wb = openpyxl.Workbook()
        plan = wb.active
        plan.title = "Plan agosto"
        plan.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        plan.append([1, "Peso muerto", 4, "8-12"])
        aux = wb.create_sheet("AUX")
        aux.append(["Nombre", "Valor"])
        aux.append(["Cumplimiento", "25%"])
        self.importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )

    def _url(self, nombre):
        return reverse(f"importaciones:{nombre}", args=[self.importacion.pk])

    def test_alumno_recibe_403(self):
        alumno = User.objects.create_user("alumno-a", password="clave-123456")
        Perfil.objects.create(
            usuario=alumno, gimnasio=self.gimnasio, rol=Perfil.Rol.ALUMNO
        )
        self.client.login(username="alumno-a", password="clave-123456")
        self.assertEqual(self.client.get(self._url("plantillas_hojas")).status_code, 403)

    def test_lista_todas_las_hojas_con_lo_que_detecto(self):
        response = self.client.get(self._url("plantillas_hojas"))
        self.assertContains(response, "Plan agosto")
        self.assertContains(response, "AUX")
        filas = {f["nombre_hoja"]: f for f in response.context["filas"]}
        self.assertEqual(filas["Plan agosto"]["cantidad"], 1)
        self.assertEqual(filas["AUX"]["cantidad"], 0)

    def test_solo_vienen_pre_marcadas_las_hojas_con_ejercicios(self):
        response = self.client.get(self._url("plantillas_hojas"))
        filas = {f["nombre_hoja"]: f for f in response.context["filas"]}
        self.assertTrue(filas["Plan agosto"]["marcada"])
        self.assertFalse(filas["AUX"]["marcada"])

    def test_la_eleccion_se_guarda_y_lleva_al_preview(self):
        response = self.client.post(
            self._url("plantillas_hojas"), {"hojas": ["Plan agosto"]}
        )
        self.assertRedirects(response, self._url("plantillas_preview"))
        self.importacion.refresh_from_db()
        self.assertEqual(self.importacion.resultado["hojas_elegidas"], ["Plan agosto"])

    def test_el_preview_solo_muestra_las_hojas_elegidas(self):
        self.client.post(self._url("plantillas_hojas"), {"hojas": ["Plan agosto"]})
        response = self.client.get(self._url("plantillas_preview"))
        nombres = [h["nombre_hoja"] for h, _ in response.context["hojas_con_form"]]
        self.assertEqual(nombres, ["Plan agosto"])

    def test_no_elegir_ninguna_no_avanza(self):
        response = self.client.post(self._url("plantillas_hojas"), {})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "al menos una hoja")
        self.importacion.refresh_from_db()
        self.assertNotIn("hojas_elegidas", self.importacion.resultado)

    def test_un_nombre_de_hoja_inventado_se_ignora(self):
        """El POST viene del cliente: se intersecta contra los nombres reales
        en vez de confiar en él, misma barrera que el re-fetch scopeado del
        resto del importador."""
        response = self.client.post(
            self._url("plantillas_hojas"), {"hojas": ["Plan agosto", "Hoja Fantasma"]}
        )
        self.assertRedirects(response, self._url("plantillas_preview"))
        self.importacion.refresh_from_db()
        self.assertEqual(self.importacion.resultado["hojas_elegidas"], ["Plan agosto"])

    def test_una_hoja_sin_ejercicios_no_se_puede_elegir(self):
        response = self.client.post(self._url("plantillas_hojas"), {"hojas": ["AUX"]})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "al menos una hoja")

    def test_una_importacion_vieja_sin_la_clave_sigue_funcionando(self):
        """Una `Importacion` EN_REVISION creada antes del deploy de esta
        pantalla no tiene `hojas_elegidas`: ahí "no eligió" significa "todas",
        que es exactamente lo que hacía antes."""
        self.assertNotIn("hojas_elegidas", self.importacion.resultado)
        response = self.client.get(self._url("plantillas_preview"))
        nombres = [h["nombre_hoja"] for h, _ in response.context["hojas_con_form"]]
        self.assertEqual(nombres, ["Plan agosto", "AUX"])

    def test_los_ejercicios_de_una_hoja_no_elegida_no_hay_que_clasificarlos(self):
        wb = openpyxl.Workbook()
        a = wb.active
        a.title = "Plan A"
        a.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        a.append([1, "Peso muerto", 4, "8"])
        b = wb.create_sheet("Plan B")
        b.append(["Dia", "Ejercicio", "Series", "Repeticiones"])
        b.append([1, "Remo con barra", 4, "8"])
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        self.client.post(
            reverse("importaciones:plantillas_hojas", args=[importacion.pk]),
            {"hojas": ["Plan A"]},
        )
        response = self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )
        pendientes = [
            f["nombre_normalizado"] for f in response.context["ejercicio_formset"].initial
        ]
        self.assertIn("peso muerto", pendientes)
        self.assertNotIn("remo con barra", pendientes)


class PreviewPlantillasMuestraLoQueEntendioTests(TestCase):
    """Con 172 items y 4 semanas, "N ejercicios detectados" no alcanza para
    verificar nada. El preview es la única defensa del entrenador contra una
    lectura mal alineada."""

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )
        self.client.login(username="staff-a", password="clave-123456")

    def _preview(self, wb):
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio, archivo=_archivo_xlsx(wb), usuario=self.usuario,
        )
        return self.client.get(
            reverse("importaciones:plantillas_preview", args=[importacion.pk])
        )

    def _wb_ancho(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plan"
        ws.append([None, " EJERCICIOS", None, "SEMANA 1", None, None, "SEMANA 2", None, None])
        ws.append([None, None, None, "Series", "Reps", "Carga", "Series", "Reps", "Carga"])
        ws.append(["DÍA 1\n• CORE", "A1.", "Plancha lateral", 4, "20", "10KG", 4, "25", "10KG"])
        ws.append([None, "A2.", "Press Pallof", 4, "20", None, 4, "25", None])
        ws.append([None, "A3.", "Remo invertido", 4, "20", None, 4, "25", None])
        return wb

    def test_muestra_la_semana_el_bloque_y_los_kilos(self):
        response = self._preview(self._wb_ancho())
        self.assertContains(response, "<th>Semana</th>", html=False)
        self.assertContains(response, "<th>Bloque</th>", html=False)
        self.assertContains(response, "<th>Kilos</th>", html=False)
        self.assertContains(response, "A1")
        self.assertContains(response, "10KG")

    def test_dice_como_leyo_la_hoja(self):
        """Si la app leyó el archivo como el layout equivocado, el entrenador
        tiene que poder verlo antes de confirmar."""
        response = self._preview(self._wb_ancho())
        self.assertContains(response, "semanas a lo ancho")
        self.assertContains(response, "fila 1")

    def test_muestra_el_nombre_del_dia(self):
        response = self._preview(self._wb_ancho())
        self.assertContains(response, "CORE")

    def test_las_filas_invalidas_se_agrupan_por_fila(self):
        """Una fila de Excel produce un item por semana: si falla en dos, se
        listaba dos veces y parecía que había el doble de problemas."""
        wb = self._wb_ancho()
        ws = wb.active
        ws.append([None, "B1.", "Sentadilla frontal", "cuatro", "20", None, "cinco", "25", None])
        response = self._preview(wb)

        agrupadas = response.context["hojas_con_form"][0][0]["invalidas_agrupadas"]
        self.assertEqual(len(agrupadas), 1)
        fila, motivos = agrupadas[0]
        self.assertEqual(fila, 6)
        self.assertEqual(len(motivos), 2)
        self.assertContains(response, "Semana 1")


class EjemploDescargableTests(TestCase):
    """El ejemplo se genera al vuelo, no es un binario versionado: un archivo
    estático se desincroniza del parser en cuanto cambian los alias y nadie se
    entera hasta que un cliente se queja.

    El test que importa es el circular: el ejemplo que le damos al entrenador
    tiene que ser un archivo que el importador sepa leer.
    """

    def setUp(self):
        self.gimnasio = Gimnasio.objects.create(nombre="Gym", slug="gym")
        self.usuario = User.objects.create_user("staff-a", password="clave-123456")
        Perfil.objects.create(
            usuario=self.usuario, gimnasio=self.gimnasio, rol=Perfil.Rol.STAFF
        )

    def test_anonimo_es_redirigido_al_login(self):
        url = reverse("importaciones:plantillas_ejemplo")
        self.assertRedirects(self.client.get(url), f"{reverse('login')}?next={url}")

    def test_alumno_recibe_403(self):
        alumno = User.objects.create_user("alumno-a", password="clave-123456")
        Perfil.objects.create(
            usuario=alumno, gimnasio=self.gimnasio, rol=Perfil.Rol.ALUMNO
        )
        self.client.login(username="alumno-a", password="clave-123456")
        response = self.client.get(reverse("importaciones:plantillas_ejemplo"))
        self.assertEqual(response.status_code, 403)

    def test_se_descarga_como_xlsx(self):
        self.client.login(username="staff-a", password="clave-123456")
        response = self.client.get(reverse("importaciones:plantillas_ejemplo"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("spreadsheetml", response["Content-Type"])
        self.assertIn("attachment", response["Content-Disposition"])

    def test_el_ejemplo_es_un_archivo_que_el_importador_sabe_leer(self):
        """Si esto falla, le estamos dando al entrenador un archivo que la app
        va a rechazar."""
        buffer = io.BytesIO()
        construir_ejemplo_plantillas().save(buffer)
        buffer.seek(0)

        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio,
            archivo=SimpleUploadedFile(
                "ejemplo.xlsx",
                buffer.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
            usuario=self.usuario,
        )

        plan = importacion.resultado["hojas"][0]
        self.assertEqual(plan["nombre_hoja"], "Mi plan")
        self.assertEqual(len(plan["items"]), 6)
        self.assertEqual(plan["filas_invalidas"], [])
        self.assertEqual(plan["dias_por_semana"], 2)
        self.assertEqual({i["semana"] for i in plan["items"]}, {1, 2})

    def test_la_hoja_de_ayuda_no_se_ofrece_como_plan(self):
        buffer = io.BytesIO()
        construir_ejemplo_plantillas().save(buffer)
        buffer.seek(0)
        importacion = previsualizar_importacion_plantillas(
            gimnasio=self.gimnasio,
            archivo=SimpleUploadedFile("ejemplo.xlsx", buffer.read()),
            usuario=self.usuario,
        )
        ayuda = importacion.resultado["hojas"][1]
        self.assertEqual(ayuda["nombre_hoja"], "Cómo llenarlo")
        self.assertEqual(ayuda["items"], [])

    def test_los_alias_que_documenta_son_los_que_el_parser_usa(self):
        """El ejemplo lee `ALIAS_PLANTILLA`, no una copia: si alguien agrega un
        sinónimo, la hoja de ayuda lo refleja sola."""
        ayuda = construir_ejemplo_plantillas()["Cómo llenarlo"]
        texto = " ".join(
            str(c.value) for fila in ayuda.iter_rows() for c in fila if c.value
        )
        for campo in ALIAS_PLANTILLA:
            self.assertIn(campo, texto)
        self.assertIn("microciclo", texto)
