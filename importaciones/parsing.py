"""Parsing puro de archivos `.xlsx` (Proyecto 2).

Este módulo NO importa nada de Django ni de los modelos de dominio
(`Ejercicio`, `RutinaPlantilla`): recibe un archivo, devuelve dataclasses.
Es lo que lo hace testeable con `SimpleTestCase` sin fixtures de tenant, y
lo que permite reusarlo desde `services.py` sin acoplar el parseo a la
persistencia. `normalizar_texto` vive acá (no en `matching.py`) porque este
módulo la necesita primero, para detectar encabezados; `matching.py` la
importa desde acá para normalizar nombres de ejercicio -- un solo lugar.
"""

import unicodedata
from dataclasses import dataclass, field

import openpyxl

ALIAS_PLANTILLA = {
    "semana": ["semana", "week", "sem"],
    "dia": ["dia", "día", "day"],
    "ejercicio": ["ejercicio", "ejercicios", "exercise", "movimiento"],
    "series": ["series", "serie", "sets"],
    "repeticiones": ["repeticiones", "reps", "repes", "rep"],
    "kilos": ["kilos", "kilogramos", "carga", "peso", "kg"],
    "descanso": ["descanso", "pausa", "rest"],
    "notas": ["notas", "nota", "observaciones", "comentarios"],
}

ALIAS_BIBLIOTECA = {
    "nombre": ["nombre", "ejercicio", "ejercicios", "exercise"],
    # "categoria" es el encabezado que usan los gimnasios reales; faltaba, y
    # por eso un Excel de 748 ejercicios entraba entero sin clasificar.
    # "músculo" NO va: `normalizar_texto` saca las tildes antes de comparar,
    # así que ese alias era código muerto -- "musculo" ya lo cubre.
    "grupo_muscular": [
        "grupo muscular",
        "grupo_muscular",
        "categoria",
        "categorias",
        "grupo",
        "musculo",
        "zona",
    ],
    "url_video": ["video", "url_video", "link", "youtube"],
}


def normalizar_texto(texto):
    """lowercase + sin tildes + espacios colapsados. `None` -> `""`."""
    if not texto:
        return ""
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFKD", str(texto))
        if not unicodedata.combining(c)
    )
    return " ".join(sin_tildes.lower().split())


class ColumnaRequeridaFaltante(Exception):
    """El archivo no tiene una columna sin la cual no se puede importar nada.

    Lleva los encabezados que SÍ se leyeron: es lo que le permite al staff
    entender que la app miró la fila equivocada (un título arriba de la
    tabla) en vez de quedarse adivinando. Antes de 2026-08-26 este caso
    devolvía una lista vacía y la app mostraba un preview de cero filas con
    el botón de confirmar habilitado, sin ningún mensaje.
    """

    def __init__(self, campo, encabezados):
        self.campo = campo
        self.encabezados = [str(e) for e in encabezados if e is not None]
        super().__init__(f"Falta la columna '{campo}'")


def _prefijo_de_palabra(alias, encabezado):
    """`alias` abre `encabezado` y termina donde termina una palabra.

    "nombre" prefija "nombre del ejercicio" pero no "nombres": el corte tiene
    que caer en un borde, o "core" matchearía "coreografia".
    """
    if not encabezado.startswith(alias):
        return False
    resto = encabezado[len(alias):]
    return resto == "" or not resto[0].isalnum()


def detectar_columnas(encabezados, alias_por_campo):
    """Devuelve (campo_canonico -> índice de columna, advertencias).

    Dos pasadas. Primero coincidencia EXACTA del encabezado normalizado
    contra la lista de alias; después, solo para los campos que quedaron sin
    columna, coincidencia por PREFIJO en borde de palabra ("nombre" al
    principio de "nombre del ejercicio").

    Es prefijo y no "contiene" a propósito. Con "contiene", una fila de
    título como "Biblioteca de ejercicios 2026" -- el caso real que dejaba
    un preview vacío -- matchea el alias "ejercicios" y se hace pasar por la
    columna de nombre, que es peor que no detectar nada. Como prefijo, esa
    fila no matchea ("biblioteca" no es alias de nada) y el archivo se
    rechaza con un error que se entiende.

    El orden importa: la parcial es útil pero ambigua -- "Grupo muscular del
    ejercicio" empieza con "grupo muscular" y también contiene "ejercicio".
    Resolver primero todo lo exacto y no reasignar columnas ya tomadas evita
    que un campo se robe la columna de otro. Entre varios alias que prefijan
    el mismo encabezado gana el más largo, por la misma razón.

    Un campo sin ninguna columna que matchee simplemente no aparece en el
    dict de salida -- el caller decide si es requerido u opcional.
    """
    normalizados = [normalizar_texto(e) for e in encabezados]
    campos = {}
    advertencias = []

    for campo, alias in alias_por_campo.items():
        indices = [i for i, valor in enumerate(normalizados) if valor in alias]
        if not indices:
            continue
        campos[campo] = indices[0]
        if len(indices) > 1:
            advertencias.append(
                f"Se encontraron {len(indices)} columnas parecidas a "
                f"'{campo}'; se usó la columna {indices[0] + 1}."
            )

    tomadas = set(campos.values())
    for campo, alias in alias_por_campo.items():
        if campo in campos:
            continue
        mejor = None  # (largo del alias, -indice de columna, indice)
        for i, valor in enumerate(normalizados):
            if i in tomadas or not valor:
                continue
            contenidos = [a for a in alias if _prefijo_de_palabra(a, valor)]
            if not contenidos:
                continue
            candidato = (len(max(contenidos, key=len)), -i, i)
            if mejor is None or candidato > mejor:
                mejor = candidato
        if mejor is not None:
            campos[campo] = mejor[2]
            tomadas.add(mejor[2])
            advertencias.append(
                f"No hay una columna llamada exactamente '{campo}'; se usó "
                f"'{encabezados[mejor[2]]}' (columna {mejor[2] + 1})."
            )

    return campos, advertencias


@dataclass(frozen=True)
class ItemParseado:
    semana: int
    dia: int
    orden: int
    ejercicio_original: str
    series: int
    repeticiones: str
    kilos: str
    descanso: str
    notas: str


@dataclass(frozen=True)
class FilaInvalida:
    fila_excel: int
    motivo: str


@dataclass(frozen=True)
class HojaParseada:
    nombre_hoja: str
    dias_por_semana: int
    items: list = field(default_factory=list)
    filas_invalidas: list = field(default_factory=list)
    # Por qué la hoja quedó con 0 items cuando falta una columna requerida
    # (`None` en el caso normal, con items). Sin esto, una hoja excluida por
    # falta de columna se ve idéntica a una hoja válida pero vacía -- el
    # staff no tiene forma de saber por qué (constraint no negociable:
    # "se excluye con motivo", fix post-review, hallazgo 2).
    motivo_exclusion: str | None = None
    # Advertencias de `detectar_columnas` (p. ej. columna duplicada) --
    # antes se calculaban y se descartaban sin llegar nunca al staff (fix
    # post-review, hallazgo 3).
    advertencias_columnas: list = field(default_factory=list)


def _mapa_merges(ws):
    """(fila, col) 1-indexed -> (fila_ancla, col_ancla) para cada celda
    dentro de un rango combinado. openpyxl devuelve `None` para toda celda
    de un merge salvo la esquina superior-izquierda; sin este mapa, una
    columna mergeada verticalmente (típico de "Semana 1" armada a mano)
    se leería como si esas filas no tuvieran valor."""
    mapa = {}
    for rango in ws.merged_cells.ranges:
        ancla = (rango.min_row, rango.min_col)
        for fila in range(rango.min_row, rango.max_row + 1):
            for col in range(rango.min_col, rango.max_col + 1):
                mapa[(fila, col)] = ancla
    return mapa


def _valor_celda(ws, fila, col, mapa_merges):
    fila_ancla, col_ancla = mapa_merges.get((fila, col), (fila, col))
    return ws.cell(row=fila_ancla, column=col_ancla).value


def _fila_vacia(valores):
    return all(v is None or str(v).strip() == "" for v in valores)


def leer_hoja_plantilla(ws):
    """Parsea una hoja de un archivo de PLANTILLAS. `ws` es una worksheet
    de `openpyxl` ya abierta (no toca el filesystem acá)."""
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    campos, advertencias = detectar_columnas(encabezados, ALIAS_PLANTILLA)

    for campo in ("ejercicio", "series", "repeticiones"):
        if campo not in campos:
            return HojaParseada(
                nombre_hoja=ws.title,
                dias_por_semana=0,
                motivo_exclusion=f"No se pudo importar: falta la columna '{campo}'",
                advertencias_columnas=advertencias,
            )

    mapa_merges = _mapa_merges(ws)
    ncols = len(encabezados)
    items = []
    filas_invalidas = []
    contador_orden = {}  # (semana, dia) -> próximo orden

    for fila_idx in range(2, ws.max_row + 1):
        valores = [_valor_celda(ws, fila_idx, c, mapa_merges) for c in range(1, ncols + 1)]
        if _fila_vacia(valores):
            continue

        ejercicio = valores[campos["ejercicio"]]
        if not ejercicio or not str(ejercicio).strip():
            filas_invalidas.append(FilaInvalida(fila_idx, "Falta el nombre del ejercicio"))
            continue

        series_raw = valores[campos["series"]]
        try:
            series = int(series_raw)
        except (TypeError, ValueError):
            filas_invalidas.append(
                FilaInvalida(fila_idx, "La columna 'series' no es un número")
            )
            continue

        repeticiones = valores[campos["repeticiones"]]
        if repeticiones is None or not str(repeticiones).strip():
            filas_invalidas.append(FilaInvalida(fila_idx, "Falta 'repeticiones'"))
            continue

        semana_raw = valores[campos["semana"]] if "semana" in campos else None
        try:
            semana = int(semana_raw) if semana_raw is not None else 1
        except (TypeError, ValueError):
            semana = 1

        dia_raw = valores[campos["dia"]] if "dia" in campos else None
        try:
            dia = int(dia_raw) if dia_raw is not None else 1
        except (TypeError, ValueError):
            filas_invalidas.append(FilaInvalida(fila_idx, "La columna 'dia' no es un número"))
            continue

        clave_orden = (semana, dia)
        contador_orden[clave_orden] = contador_orden.get(clave_orden, 0) + 1

        kilos = valores[campos["kilos"]] if "kilos" in campos else None
        descanso = valores[campos["descanso"]] if "descanso" in campos else None
        notas = valores[campos["notas"]] if "notas" in campos else None

        items.append(ItemParseado(
            semana=semana,
            dia=dia,
            orden=contador_orden[clave_orden],
            ejercicio_original=str(ejercicio).strip(),
            series=series,
            repeticiones=str(repeticiones).strip(),
            kilos=str(kilos).strip() if kilos else "",
            descanso=str(descanso).strip() if descanso else "",
            notas=str(notas).strip() if notas else "",
        ))

    dias_por_semana = max((i.dia for i in items), default=0)
    return HojaParseada(
        nombre_hoja=ws.title,
        dias_por_semana=dias_por_semana,
        items=items,
        filas_invalidas=filas_invalidas,
        advertencias_columnas=advertencias,
    )


def leer_hoja_biblioteca(ws):
    """Parsea una hoja del import de BIBLIOTECA: solo nombre + grupo
    muscular (opcional) + video (opcional), sin días/semanas/series."""
    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    campos, advertencias = detectar_columnas(encabezados, ALIAS_BIBLIOTECA)

    if "nombre" not in campos:
        # Antes esto devolvía `[], [], advertencias` y la app mostraba un
        # preview vacío con el botón de confirmar habilitado. Sin el nombre
        # del ejercicio no hay nada que importar: es un error, no un archivo
        # de cero filas.
        raise ColumnaRequeridaFaltante("nombre", encabezados)

    mapa_merges = _mapa_merges(ws)
    ncols = len(encabezados)
    items = []
    filas_invalidas = []

    for fila_idx in range(2, ws.max_row + 1):
        valores = [_valor_celda(ws, fila_idx, c, mapa_merges) for c in range(1, ncols + 1)]
        if _fila_vacia(valores):
            continue

        nombre = valores[campos["nombre"]]
        if not nombre or not str(nombre).strip():
            filas_invalidas.append(FilaInvalida(fila_idx, "Falta el nombre del ejercicio"))
            continue

        grupo_muscular = valores[campos["grupo_muscular"]] if "grupo_muscular" in campos else None
        url_video = valores[campos["url_video"]] if "url_video" in campos else None

        items.append({
            "fila_excel": fila_idx,
            "nombre_original": str(nombre).strip(),
            "grupo_muscular_original": str(grupo_muscular).strip() if grupo_muscular else None,
            "url_video": str(url_video).strip() if url_video else "",
        })

    return items, filas_invalidas, advertencias


def parsear_archivo_plantillas(archivo):
    """Abre `archivo` (un `UploadedFile` de Django) y devuelve una
    `HojaParseada` por cada hoja del workbook (decisión 7 del spec:
    multi-hoja -> multi-plantilla)."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    return [leer_hoja_plantilla(wb[nombre]) for nombre in wb.sheetnames]


def parsear_archivo_biblioteca(archivo):
    """El import de biblioteca usa solo la primera hoja del archivo."""
    wb = openpyxl.load_workbook(archivo, data_only=True)
    return leer_hoja_biblioteca(wb[wb.sheetnames[0]])
